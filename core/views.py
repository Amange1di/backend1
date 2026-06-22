from datetime import date, timedelta
from calendar import monthrange
import re

from django.db import models
from django.db.models import Sum
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import ensure_csrf_cookie
from django.middleware.csrf import get_token
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.authtoken.models import Token
from rest_framework.parsers import FormParser, MultiPartParser, JSONParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.throttling import AnonRateThrottle
from django.conf import settings
from django.template.loader import render_to_string
from django.http import HttpResponse
import io
import os

import logging
import bleach

from .models import (
    Attendance,
    Auditorium,
    Course,
    Expense,
    Group,
    GroupMonth,
    LandingHeaderLink,
    LandingPage,
    LandingSection,
    HomeworkTaskAttachment,
    HomeworkSubmission,
    HomeworkTask,
    Payment,
    Student,
    TrialLead,
    Task,
    TaskLead,
    User,
    TelegramBindCode,
    CompanyBalance,
    PromoCode,
    Transaction,
    Company,
    CompanyCategory,
    CompanyCity,
    PublicCourse,
    JobVacancy,
    StudentApplication,
    UserBalance,
    Contract,
    ContractTemplate,
)
from .permissions import (
    IsAdmin,
    IsCourseAdmin,
    IsCourseAdminOrManager,
    IsCourseAdminOrManagerReadOnly,
    IsCourseAdminOrManagerOrStudentReadOnly,
    IsCourseAdminOrTeacherReadOnly,
    IsTeacherOrCourseAdminReadOnly,
)

logger = logging.getLogger(__name__)


class LoginThrottle(AnonRateThrottle):
    """Rate limiting для login endpoints"""
    rate = '1000/hour'


class RegisterThrottle(AnonRateThrottle):
    """Rate limiting для registration endpoints"""
    rate = '1000/hour'


class PublicSubmitThrottle(AnonRateThrottle):
    """Rate limiting для публичных форм отправки (landing leads, crm contact)"""
    rate = '10/minute'


class PublicReadThrottle(AnonRateThrottle):
    """Rate limiting для публичных GET endpoints (landing pages)"""
    rate = '60/minute'


from .permissions import (
    IsAdmin,
    IsCourseAdmin,
    IsCourseAdminOrManager,
    IsCourseAdminOrManagerReadOnly,
    IsCourseAdminOrManagerOrStudentReadOnly,
    IsCourseAdminOrTeacherReadOnly,
    IsTeacherOrCourseAdminReadOnly,
)
from .serializers import (
    AttendanceSerializer,
    AuditoriumSerializer,
    CourseAdminUpdateSerializer,
    CourseSerializer,
    ExpenseSerializer,
    GroupMonthSerializer,
    GroupSerializer,
    LandingHeaderLinkSerializer,
    LandingPageSerializer,
    LandingPublicPageSerializer,
    HomeworkSubmissionSerializer,
    HomeworkTaskSerializer,
    LoginSerializer,
    PaymentSerializer,
    RegisterSerializer,
    StudentIdentityLoginSerializer,
    StudentProfileSerializer,
    StudentSetPasswordSerializer,
    StudentSerializer,
    TeacherCreateSerializer,
    TeacherUpdateSerializer,
    TransferGroupSerializer,
    TrialLeadSerializer,
    TaskSerializer,
    UserUpdateSerializer,
    UserSerializer,
    PromoCodeSerializer,
    CompanySerializer,
    PublicCourseSerializer,
    JobVacancySerializer,
    CompanyCreateUpdateSerializer,
    normalize_phone,
    sync_student_user,
    ContractSerializer,
    ContractTemplateSerializer,
)


class RegisterView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        if request.user.is_authenticated and request.user.role == User.Role.ADMIN:
            return Response(
                {"detail": "Admins can only create course admins."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if request.user.is_authenticated and request.user.role == User.Role.TEACHER:
            return Response(
                {"detail": "Teachers cannot create users."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if request.user.is_authenticated and request.user.role == User.Role.MANAGER:
            return Response(
                {"detail": "Managers cannot create users."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if request.user.is_authenticated and request.user.role == User.Role.STUDENT:
            return Response(
                {"detail": "Students cannot create users."},
                status=status.HTTP_403_FORBIDDEN,
            )
        serializer = RegisterSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        requested_role = serializer.validated_data.get("role")
        if (
            (not request.user.is_authenticated
             or request.user.role != User.Role.COURSE_ADMIN)
            and requested_role == User.Role.MANAGER
        ):
            return Response(
                {"detail": "Managers can only be created by course admins."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if (
            request.user.is_authenticated
            and request.user.role == User.Role.COURSE_ADMIN
        ):
            if requested_role == User.Role.MANAGER:
                # Check if course admin can create more managers
                if not request.user.can_create_manager():
                    return Response(
                        {
                            "detail": f"Manager limit reached. Maximum: {request.user.max_managers}, "
                            f"Current: {request.user.get_managers_count()}"
                        },
                        status=status.HTTP_403_FORBIDDEN,
                    )
                user = serializer.save(
                    force_role=User.Role.MANAGER,
                    created_by=request.user,
                )
            else:
                user = serializer.save(
                    force_role=User.Role.TEACHER,
                    created_by=request.user,
                )
        else:
            user = serializer.save(force_role=User.Role.TEACHER)
        token, _ = Token.objects.get_or_create(user=user)
        return Response(
            {"token": token.key, "user": UserSerializer(user).data},
            status=status.HTTP_201_CREATED,
        )


@method_decorator(ensure_csrf_cookie, name='dispatch')
class LoginView(APIView):
    permission_classes = [permissions.AllowAny]
    throttle_classes = [LoginThrottle]

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data["user"]
        token, _ = Token.objects.get_or_create(user=user)
        
        # Устанавливаем CSRF cookie явно
        response = Response({"token": token.key, "user": UserSerializer(user).data})
        response.set_cookie(
            key="csrftoken",
            value=get_token(request),
            max_age=60 * 60 * 24 * 30,  # 30 дней
            httponly=False,  # JS может читать для формы
            samesite="Lax",
            secure=settings.DEBUG is False,
        )
        
        # httpOnly cookie с токеном — защита от XSS кражи токена
        # JS не может прочитать этот cookie, только браузер отправляет его автоматически
        response.set_cookie(
            key="token",
            value=token.key,
            max_age=60 * 60 * 24 * 30,  # 30 дней
            httponly=True,  # КЛЮЧЕВОЕ: JS не может прочитать
            samesite="Lax",
            secure=not settings.DEBUG,
            path="/",
        )
        return response


class LogoutView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        """Выход из системы - удаление токена"""
        try:
            request.user.auth_token.delete()
        except Exception:
            pass
        
        response = Response({"detail": "Successfully logged out."})
        # Очищаем CSRF cookie и httpOnly token cookie
        response.delete_cookie("csrftoken")
        response.delete_cookie("token", path="/")
        return response


class StudentLoginView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = StudentIdentityLoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        phone_number = serializer.validated_data["phone_number"].strip()
        password = serializer.validated_data.get("password", "")
        normalized_phone = normalize_phone(phone_number)

        candidates = [
            student
            for student in Student.objects.select_related("user")
            .all()
            .order_by("id")
            if normalize_phone(student.phone) == normalized_phone
        ]

        if not candidates:
            return Response(
                {"detail": "Invalid student credentials."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        accessible_students = []
        for student in candidates:
            if not student.user:
                sync_student_user(student)
                student.refresh_from_db()
            try:
                ensure_student_access_allowed(student)
                accessible_students.append(student)
            except PermissionDenied:
                continue

        if not accessible_students:
            return Response(
                {"detail": "Student access is disabled."},
                status=status.HTTP_403_FORBIDDEN,
            )

        if len(accessible_students) > 1:
            return Response(
                {"detail": "Multiple student accounts matched. Contact your administrator."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        student = accessible_students[0]
        user = student.user
        token, _ = Token.objects.get_or_create(user=user)

        if user.must_set_password or not user.has_usable_password():
            return Response(
                {
                    "token": token.key,
                    "user": UserSerializer(user).data,
                    "student": StudentSerializer(student, context={"request": request}).data,
                    "requires_password_setup": True,
                }
            )

        if not password:
            return Response(
                {
                    "detail": "Password is required.",
                    "code": "password_required",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not user.check_password(password):
            return Response(
                {"detail": "Invalid student credentials."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {
                "token": token.key,
                "user": UserSerializer(user).data,
                "student": StudentSerializer(student, context={"request": request}).data,
                "requires_password_setup": False,
            }
        )


class StudentSetPasswordView(APIView):
    def post(self, request):
        if request.user.role != User.Role.STUDENT:
            raise PermissionDenied("Only students can set this password.")
        serializer = StudentSetPasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        request.user.set_password(serializer.validated_data["password"])
        request.user.must_set_password = False
        request.user.save(update_fields=["password", "must_set_password"])
        token, _ = Token.objects.get_or_create(user=request.user)
        return Response({"token": token.key, "user": UserSerializer(request.user).data})


class StudentProfileView(APIView):
    def get(self, request):
        if request.user.role != User.Role.STUDENT:
            raise PermissionDenied("Only students can access this profile.")
        student = getattr(request.user, "student_profile", None)
        if not student:
            return Response({"detail": "Student profile not found."}, status=status.HTTP_404_NOT_FOUND)
        ensure_student_access_allowed(student)
        return Response(
            {
                "id": student.id,
                "first_name": student.first_name,
                "last_name": student.last_name,
                "phone": student.phone,
                "telegram": student.telegram,
                "company_name": student.company.name if student.company else "",
                "can_login": student.can_login,
                "must_set_password": request.user.must_set_password,
            }
        )

    def patch(self, request):
        if request.user.role != User.Role.STUDENT:
            raise PermissionDenied("Only students can update this profile.")
        student = getattr(request.user, "student_profile", None)
        if not student:
            return Response({"detail": "Student profile not found."}, status=status.HTTP_404_NOT_FOUND)
        ensure_student_access_allowed(student)
        serializer = StudentProfileSerializer(data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)

        student_fields = []
        user_fields = []
        phone = serializer.validated_data.get("phone")
        telegram = serializer.validated_data.get("telegram")
        password = serializer.validated_data.get("password")

        if phone is not None and phone != student.phone:
            student.phone = phone
            request.user.phone = phone
            student_fields.append("phone")
            user_fields.append("phone")
        if telegram is not None and telegram != student.telegram:
            student.telegram = telegram
            request.user.telegram = telegram
            student_fields.append("telegram")
            user_fields.append("telegram")
        if student_fields:
            student.save(update_fields=student_fields)
        if password:
            request.user.set_password(password)
            request.user.must_set_password = False
            user_fields.extend(["password", "must_set_password"])
        if user_fields:
            request.user.save(update_fields=list(dict.fromkeys(user_fields)))

        return self.get(request)


class CourseAdminCreateView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        admins = User.objects.filter(role=User.Role.COURSE_ADMIN).order_by(
            "-date_joined"
        )
        return Response(UserSerializer(admins, many=True).data)

    def post(self, request):
        serializer = RegisterSerializer(
            data=request.data, context={"force_role": User.Role.COURSE_ADMIN}
        )
        serializer.is_valid(raise_exception=True)
        company_name = serializer.validated_data.get("company_name", "").strip()
        phone = serializer.validated_data.get("phone", "").strip()
        address = serializer.validated_data.get("address", "").strip()
        max_managers = serializer.validated_data.get("max_managers", 0) or 0
        if not company_name:
            return Response(
                {"detail": "Company name is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not phone or not address:
            return Response(
                {"detail": "Phone and address are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if max_managers < 0:
            return Response(
                {"detail": "Manager limit must be 0 or more."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        # Сначала создаём пользователя без компании
        user = serializer.save(
            created_by=request.user,
            company=None,  # Сначала без компании
            max_managers=max_managers,
        )
        
        # Теперь создаём компанию с этим пользователем как owner
        company = Company.objects.create(
            name=company_name,
            owner=user,
            category=CompanyCategory.OTHER,
            city=CompanyCity.ONLINE,
            description=f"Company for {company_name}",
            is_active=True,
        )
        
        # Обновляем пользователя, связывая с компанией
        user.company = company
        user.save()
        
        # Создаём баланс для компании
        CompanyBalance.objects.create(company=company, balance=0)
        
        token, _ = Token.objects.get_or_create(user=user)
        return Response(
            {"token": token.key, "user": UserSerializer(user).data},
            status=status.HTTP_201_CREATED,
        )


class CourseAdminDetailView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request, pk: int):
        admin = get_object_or_404(User, pk=pk, role=User.Role.COURSE_ADMIN)
        return Response(UserSerializer(admin).data)

    def patch(self, request, pk: int):
        admin = get_object_or_404(User, pk=pk, role=User.Role.COURSE_ADMIN)
        serializer = CourseAdminUpdateSerializer(admin, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        admin = serializer.save()
        return Response(UserSerializer(admin).data)

    def delete(self, request, pk: int):
        admin = get_object_or_404(User, pk=pk, role=User.Role.COURSE_ADMIN)
        admin.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MeView(APIView):
    def get(self, request):
        data = UserSerializer(request.user).data
        if request.user.is_superuser and data.get("role") != User.Role.ADMIN:
            data["role"] = User.Role.ADMIN
        if request.user.role == User.Role.STUDENT:
            data["student_id"] = (
                request.user.student_profile.id
                if hasattr(request.user, "student_profile")
                and request.user.student_profile
                else None
            )
        return Response(
            {
                **data,
                "support_telegram": resolve_support_telegram(request.user),
            }
        )


def resolve_user_company_name(user: User) -> str:
    if getattr(user, "company", None) and user.company.name:
        return user.company.name.strip()
    if user.role == User.Role.MANAGER and user.created_by:
        if getattr(user.created_by, "company", None) and user.created_by.company.name:
            return user.created_by.company.name.strip()
    return ""


class UserBalanceHistoryView(APIView):
    """История транзакций eduCoin для компании"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            return Response(
                {"detail": "Доступно только для course_admin и manager."},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        company_name = resolve_user_company_name(request.user)
        if not company_name:
            return Response(
                {"detail": "Компания не найдена."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        # Находим компанию пользователя (учитывая created_by для менеджеров)
        user_company = request.user.company
        if not user_company and request.user.role == User.Role.MANAGER:
            company_name = resolve_user_company_name(request.user)
            if company_name:
                user_company = Company.objects.filter(name=company_name).first()
        if not user_company:
            return Response(
                {"detail": "Компания не найдена."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        transactions = Transaction.objects.filter(
            company=user_company
        ).order_by("-timestamp")
        
        # Получаем баланс компании
        balance = 0
        try:
            company_balance = CompanyBalance.objects.get(company=user_company)
            balance = company_balance.balance
        except CompanyBalance.DoesNotExist:
            pass
        
        data = []
        for t in transactions:
            data.append({
                "id": t.id,
                "amount": t.amount,
                "reason": t.reason,
                "transaction_type": t.transaction_type,
                "transaction_type_display": t.get_transaction_type_display(),
                "timestamp": t.timestamp.isoformat(),
                "balance_after": balance,
            })
            balance -= t.amount
        
        return Response({
            "balance": CompanyBalance.objects.filter(company=user_company).first().balance if CompanyBalance.objects.filter(company=user_company).exists() else 0,
            "transactions": data,
        })


class UserBalanceMeView(APIView):
    """Текущий баланс компании eduCoin"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            return Response(
                {"detail": "Доступно только для course_admin и manager."},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        # Находим компанию пользователя (учитывая created_by для менеджеров)
        user_company = request.user.company
        if not user_company and request.user.role == User.Role.MANAGER:
            company_name = resolve_user_company_name(request.user)
            if company_name:
                user_company = Company.objects.filter(name=company_name).first()
        if not user_company:
            return Response(
                {"balance": 0},
            )
        
        balance = 0
        try:
            company_balance = CompanyBalance.objects.get(company=user_company)
            balance = company_balance.balance
        except CompanyBalance.DoesNotExist:
            pass
        
        return Response({
            "balance": balance,
            "company_name": user_company.name,
        })


def resolve_support_telegram(user: User) -> str:
    if user.role == User.Role.COURSE_ADMIN:
        admin = (
            user.created_by
            if user.created_by and user.created_by.role == User.Role.ADMIN
            else None
        )
        if admin and admin.telegram:
            return admin.telegram
        admin = (
            User.objects.filter(role=User.Role.ADMIN).order_by("date_joined").first()
        )
        return admin.telegram if admin and admin.telegram else ""
    if user.role in (User.Role.TEACHER, User.Role.MANAGER, User.Role.STUDENT):
        if user.created_by and user.created_by.telegram:
            return user.created_by.telegram
        if user.company:
            course_admin = (
                User.objects.filter(
                    role=User.Role.COURSE_ADMIN, company=user.company
                )
                .order_by("date_joined")
                .first()
            )
            return (
                course_admin.telegram if course_admin and course_admin.telegram else ""
            )
        return ""
    if user.role == User.Role.ADMIN:
        return user.telegram or ""
    return ""


def get_company_student_cabinet_enabled(company):
    if not company:
        return False
    return User.objects.filter(
        role=User.Role.COURSE_ADMIN,
        company=company,
        is_student_cabinet_enabled=True,
    ).exists()


def student_has_allowed_group(student: Student) -> bool:
    groups = student.groups.all()
    if not groups.exists():
        return True
    return groups.filter(is_login_allowed=True).exists()


def ensure_student_access_allowed(student: Student):
    if not student.company or not get_company_student_cabinet_enabled(student.company.name if student.company else ""):
        raise PermissionDenied("Student cabinet is disabled for this company.")
    if not student.can_login:
        raise PermissionDenied("Student login is disabled for this account.")
    if not student_has_allowed_group(student):
        raise PermissionDenied("Student login is disabled for this group.")
    if not student.user or student.user.role != User.Role.STUDENT:
        raise PermissionDenied("Student account is not configured.")
    if not student.user.is_active:
        raise PermissionDenied("Student account is inactive.")


def _student_can_access_homework_task(task: HomeworkTask, student: Student) -> bool:
    if task.target_type == HomeworkTask.TargetType.SPECIFIC_STUDENTS:
        return task.students.filter(id=student.id).exists()
    return task.group.students.filter(id=student.id).exists()


def _is_submission_locked(task: HomeworkTask) -> bool:
    if not task.hard_deadline:
        return False
    grace_delta = timedelta(minutes=task.grace_period_minutes or 0)
    if task.allow_late:
        return False
    return timezone.now() > (task.deadline + grace_delta)


def parse_schedule_days(value: str) -> set[int]:
    normalized = (
        value.lower()
        .replace(".", " ")
        .replace(",", " ")
        .replace(";", " ")
        .replace("/", " ")
    )
    tokens = [token for token in normalized.split() if token]
    mapping = [
        (
            0,
            [
                "mon",
                "monday",
                "\u043f\u043d",
                "\u0434\u04af\u0439",
                "\u0434\u04af\u0439\u0448",
            ],
        ),
        (1, ["tue", "tuesday", "\u0432\u0442", "\u0448\u0435\u0439"]),
        (2, ["wed", "wednesday", "\u0441\u0440", "\u0448\u0430\u0440"]),
        (3, ["thu", "thursday", "\u0447\u0442", "\u0431\u0435\u0439"]),
        (4, ["fri", "friday", "\u043f\u0442", "\u0436\u0443\u043c"]),
        (5, ["sat", "saturday", "\u0441\u0431", "\u0438\u0448"]),
        (6, ["sun", "sunday", "\u0432\u0441", "\u0436\u0435\u043a"]),
    ]
    result: set[int] = set()
    for token in tokens:
        for idx, keys in mapping:
            if any(token.startswith(key) for key in keys):
                result.add(idx)
                break
    return result


def parse_time_to_minutes(value: str):
    if not value:
        return None
    match = re.match(r"^(\d{1,2})[:.](\d{2})$", value.strip())
    if not match:
        return None
    hours = int(match.group(1))
    minutes = int(match.group(2))
    if hours < 0 or hours > 23 or minutes < 0 or minutes > 59:
        return None
    return hours * 60 + minutes


def ranges_overlap(start_a, end_a, start_b, end_b):
    a_start = start_a or date(1970, 1, 1)
    a_end = end_a or date(2999, 12, 31)
    b_start = start_b or date(1970, 1, 1)
    b_end = end_b or date(2999, 12, 31)
    return a_start <= b_end and b_start <= a_end


def time_ranges_overlap(start_a: int, end_a: int, start_b: int, end_b: int) -> bool:
    return start_a < end_b and start_b < end_a


def compute_group_end_date(start_date, schedule_days: str, lessons_count):
    if not start_date or not lessons_count:
        return None
    if not schedule_days:
        return None
    days_set = parse_schedule_days(schedule_days)
    if not days_set:
        return None
    total = int(lessons_count)
    if total <= 0:
        return None
    current = start_date
    count = 0
    while count < total:
        if current.weekday() in days_set:
            count += 1
            if count == total:
                break
        current = current + timedelta(days=1)
    return current


class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.all().order_by("-created_at")
    serializer_class = CourseSerializer
    permission_classes = [IsCourseAdminOrManagerReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == User.Role.COURSE_ADMIN:
            return queryset.filter(admins=user)
        if user.is_authenticated and user.role == User.Role.MANAGER:
            if not user.company:
                return queryset.none()
            return queryset.filter(admins__company=user.company).distinct()
        return queryset

    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == User.Role.MANAGER:
            raise PermissionDenied("Managers cannot create courses.")
        admins = serializer.validated_data.get("admins", [])
        if user.role == User.Role.COURSE_ADMIN:
            for admin in admins:
                if admin.company != user.company:
                    raise PermissionDenied(
                        "Course admins can only assign their company admins."
                    )
            course = serializer.save()
            course.admins.add(user)
            if admins:
                course.admins.add(*admins)
            return
        serializer.save()

    def perform_update(self, serializer):
        user = self.request.user
        if user.role == User.Role.MANAGER:
            raise PermissionDenied("Managers cannot update courses.")
        if user.role == User.Role.COURSE_ADMIN:
            admins = serializer.validated_data.get("admins", None)
            if admins is not None:
                for admin in admins:
                    if admin.company != user.company:
                        raise PermissionDenied(
                            "Course admins can only assign their company admins."
                        )
            course = serializer.save()
            course.admins.add(user)
            return
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        if request.user.role == User.Role.MANAGER:
            raise PermissionDenied("Managers cannot delete courses.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["get"])
    def stats(self, request, pk=None):
        course = self.get_object()
        students_qs = Student.objects.filter(primary_course=course)
        students_count = students_qs.count()
        paid_students = (
            Payment.objects.filter(status=Payment.Status.PAID, student__in=students_qs)
            .values("student")
            .distinct()
            .count()
        )
        attendance_qs = Attendance.objects.filter(group__course=course)
        total_attendance = attendance_qs.count()
        present_count = attendance_qs.filter(status=Attendance.Status.PRESENT).count()
        excused_count = attendance_qs.filter(status=Attendance.Status.EXCUSED).count()
        absent_count = attendance_qs.filter(status=Attendance.Status.ABSENT).count()
        attendance_rate = (
            (present_count + excused_count) / total_attendance
            if total_attendance
            else 0
        )
        return Response(
            {
                "students_total": students_count,
                "students_paid": paid_students,
                "attendance_total": total_attendance,
                "attendance_present": present_count,
                "attendance_excused": excused_count,
                "attendance_absent": absent_count,
                "attendance_rate": round(attendance_rate, 4),
            }
        )


class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all().order_by("-created_at")
    serializer_class = StudentSerializer
    permission_classes = [IsCourseAdminOrManagerOrStudentReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == User.Role.COURSE_ADMIN:
            return queryset.filter(
                models.Q(primary_course__admins=user)
                | models.Q(groups__course__admins=user)
            ).distinct()
        if user.is_authenticated and user.role == User.Role.MANAGER:
            if not user.company:
                return queryset.none()
            return queryset.filter(
                models.Q(company=user.company)
                | models.Q(primary_course__admins=user)
                | models.Q(groups__company=user.company)
                | models.Q(groups__course__admins=user)
            ).distinct()
        if user.is_authenticated and user.role == User.Role.STUDENT:
            return queryset.filter(user=user)
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        if user.role in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            course = serializer.validated_data.get("primary_course")
            group_ids = serializer.validated_data.get("group_ids", [])
            account = serializer.validated_data.get("user")
            auto_course = None
            if course:
                allowed = course.admins.filter(id=user.id).exists()
                if user.role == User.Role.MANAGER:
                    allowed = course.admins.filter(
                        company=user.company
                    ).exists()
                if not allowed:
                    raise PermissionDenied("Not allowed for this course.")
            for group in group_ids:
                if group.course:
                    allowed = group.course.admins.filter(id=user.id).exists()
                    if user.role == User.Role.MANAGER:
                        allowed = group.course.admins.filter(
                            company=user.company
                        ).exists()
                    if not allowed:
                        raise PermissionDenied("Not allowed for this group.")
                elif group.company and group.company != user.company:
                    raise PermissionDenied("Not allowed for this group.")
            if not course and group_ids:
                first_course = group_ids[0].course
                if first_course and all(
                    group.course_id == first_course.id for group in group_ids
                ):
                    auto_course = first_course
            if (
                account
                and account.company
                and account.company != user.company
            ):
                raise PermissionDenied("Not allowed for this user.")
            save_kwargs = {"company": user.company}
            if course:
                save_kwargs["primary_course"] = course
            elif auto_course:
                save_kwargs["primary_course"] = auto_course
            student = serializer.save(**save_kwargs)
            sync_student_user(student, created_by=user)
            return
        student = serializer.save()
        sync_student_user(student, created_by=user if user.is_authenticated else None)

    def perform_update(self, serializer):
        user = self.request.user
        if user.role in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            if user.role == User.Role.MANAGER and "can_login" in serializer.validated_data:
                raise PermissionDenied("Managers cannot change student login access.")
            course = serializer.validated_data.get("primary_course", None)
            if course:
                allowed = course.admins.filter(id=user.id).exists()
                if user.role == User.Role.MANAGER:
                    allowed = course.admins.filter(
                        company=user.company
                    ).exists()
                if not allowed:
                    raise PermissionDenied("Not allowed for this course.")
            group_ids = serializer.validated_data.get("group_ids", [])
            for group in group_ids:
                if group.course:
                    allowed = group.course.admins.filter(id=user.id).exists()
                    if user.role == User.Role.MANAGER:
                        allowed = group.course.admins.filter(
                            company=user.company
                        ).exists()
                    if not allowed:
                        raise PermissionDenied("Not allowed for this group.")
                elif group.company and group.company != user.company:
                    raise PermissionDenied("Not allowed for this group.")
            auto_course = None
            if not course and group_ids:
                first_course = group_ids[0].course
                if first_course and all(
                    group.course_id == first_course.id for group in group_ids
                ):
                    auto_course = first_course
            save_kwargs = {}
            if course:
                save_kwargs["primary_course"] = course
            elif auto_course:
                save_kwargs["primary_course"] = auto_course
            student = serializer.save(**save_kwargs)
            sync_student_user(student, created_by=user)
            return
        student = serializer.save()
        sync_student_user(student, created_by=user if user.is_authenticated else None)

    def destroy(self, request, *args, **kwargs):
        if request.user.role in (User.Role.MANAGER, User.Role.STUDENT):
            raise PermissionDenied("Not allowed to delete students.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["post"], url_path="reset-password")
    def reset_password(self, request, pk=None):
        if request.user.role != User.Role.COURSE_ADMIN:
            raise PermissionDenied("Only course admins can reset student passwords.")
        student = self.get_object()
        if student.company != request.user.company:
            raise PermissionDenied("Not allowed for this student.")
        if not student.user:
            sync_student_user(student, created_by=request.user)
            student.refresh_from_db()
        student.user.set_unusable_password()
        student.user.must_set_password = True
        student.user.save(update_fields=["password", "must_set_password"])
        Token.objects.filter(user=student.user).delete()
        return Response({"detail": "Student password was reset.", "must_set_password": True})

    @action(detail=True, methods=["post"], url_path="transfer-group")
    def transfer_group(self, request, pk=None):
        """
        Transfer a student to a new group while preserving all legacy data
        (attendance, payments, homework submissions) in the old group.
        """
        student = self.get_object()
        serializer = TransferGroupSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        new_group = serializer.validated_data["new_group"]
        note = serializer.validated_data.get("note", "")

        # Permission check
        user = request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            raise PermissionDenied("Only course admins and managers can transfer students.")

        # Ensure student and group belong to the same company
        if student.company != new_group.company:
            raise PermissionDenied("Student and new group must belong to the same company.")

        # Check company access
        if user.role == User.Role.COURSE_ADMIN and student.company != user.company:
            raise PermissionDenied("Not allowed for this student.")
        if user.role == User.Role.COURSE_ADMIN and new_group.company != user.company:
            raise PermissionDenied("Not allowed for this group.")
        if user.role == User.Role.MANAGER and student.company != user.company:
            raise PermissionDenied("Not allowed for this student.")
        if user.role == User.Role.MANAGER and new_group.company != user.company:
            raise PermissionDenied("Not allowed for this group.")

        # Unenroll from old group — history stays (FK relationships preserved)
        student.groups.clear()

        # Enroll in new group
        student.groups.add(new_group)

        # Update primary_course to match new group's course
        if new_group.course:
            student.primary_course = new_group.course

        # Add note about transfer
        if note:
            existing = student.notes or ""
            timestamp = timezone.now().strftime("%d.%m.%Y %H:%M")
            transfer_note = f"[{timestamp}] Переведён в группу «{new_group.name}». {note}"
            student.notes = f"{transfer_note}\n{existing}" if existing else transfer_note

        student.save(update_fields=["primary_course", "notes"])

        return Response({
            "status": "ok",
            "detail": f"Student transferred to group «{new_group.name}».",
        })


class TeacherViewSet(viewsets.ModelViewSet):
    queryset = User.objects.filter(role=User.Role.TEACHER).order_by("-date_joined")
    serializer_class = UserSerializer
    permission_classes = [IsCourseAdminOrManagerReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset().prefetch_related("teaching_courses")
        user = self.request.user
        if user.is_authenticated and user.role in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            if user.company:
                queryset = queryset.filter(company=user.company)
            else:
                queryset = queryset.none()
        course_param = self.request.query_params.get("course")
        if course_param:
            try:
                course_id = int(course_param)
            except (TypeError, ValueError):
                return queryset.none()
            queryset = queryset.filter(teaching_courses__id=course_id)
        return queryset.distinct()

    def create(self, request, *args, **kwargs):
        user = request.user
        if user.role == User.Role.ADMIN:
            raise PermissionDenied("Admins cannot create teachers.")
        if user.role == User.Role.MANAGER:
            raise PermissionDenied("Managers cannot create teachers.")
        serializer = TeacherCreateSerializer(
            data=request.data, context={"request": request}
        )
        serializer.is_valid(raise_exception=True)
        teacher = serializer.save()
        return Response(
            UserSerializer(teacher).data,
            status=status.HTTP_201_CREATED,
        )

    def perform_update(self, serializer):
        user = self.request.user
        if user.role == User.Role.MANAGER:
            raise PermissionDenied("Managers cannot update teachers.")
        if user.role == User.Role.COURSE_ADMIN:
            if (
                "role" in serializer.validated_data
                and serializer.validated_data["role"] != User.Role.TEACHER
            ):
                raise PermissionDenied("Course admins cannot change roles.")
        if (
            "company" in serializer.validated_data
            and serializer.validated_data.get("company") != user.company
        ):
            raise PermissionDenied("Not allowed for this company.")
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        if request.user.role == User.Role.MANAGER:
            raise PermissionDenied("Managers cannot delete teachers.")
        return super().destroy(request, *args, **kwargs)

    def get_serializer_class(self):
        if self.action in ["update", "partial_update"]:
            return TeacherUpdateSerializer
        return super().get_serializer_class()


class ManagerViewSet(viewsets.ModelViewSet):
    queryset = User.objects.filter(role=User.Role.MANAGER).order_by("-date_joined")
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == User.Role.COURSE_ADMIN:
            if user.company:
                return queryset.filter(company=user.company)
            return queryset.none()
        if user.is_authenticated and user.role == User.Role.MANAGER:
            if user.company:
                return queryset.filter(company=user.company)
            return queryset.none()
        return queryset.none()

    def create(self, request, *args, **kwargs):
        user = request.user
        if user.role != User.Role.COURSE_ADMIN:
            raise PermissionDenied("Only course admins can create managers.")
        if not user.can_create_manager():
            raise PermissionDenied(
                f"Manager limit reached. Maximum: {user.max_managers}, "
                f"Current: {user.get_managers_count()}"
            )
        serializer = RegisterSerializer(
            data=request.data, context={"force_role": User.Role.MANAGER}
        )
        serializer.is_valid(raise_exception=True)
        manager = serializer.save(created_by=user, company=user.company)
        return Response(UserSerializer(manager).data, status=status.HTTP_201_CREATED)

    def get_serializer_class(self):
        if self.action in ["update", "partial_update"]:
            return UserUpdateSerializer
        return super().get_serializer_class()


class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all().order_by("-created_at")
    serializer_class = GroupSerializer
    permission_classes = [IsCourseAdminOrTeacherReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == User.Role.COURSE_ADMIN:
            return queryset.filter(
                models.Q(course__admins=user) | models.Q(company=user.company)
            ).distinct()
        if user.is_authenticated and user.role == User.Role.MANAGER:
            return queryset.filter(
                models.Q(course__admins__company=user.company)
                | models.Q(company=user.company)
            ).distinct()
        if user.is_authenticated and user.role == User.Role.TEACHER:
            return queryset.filter(teacher=user)
        if user.is_authenticated and user.role == User.Role.STUDENT:
            return queryset.filter(students__user=user).distinct()
        return queryset

    def create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        user = self.request.user
        if user.role in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            course = serializer.validated_data.get("course")
            if not course:
                raise PermissionDenied("Not allowed for this course.")
            allowed = course.admins.filter(id=user.id).exists()
            if user.role == User.Role.MANAGER:
                allowed = course.admins.filter(company=user.company).exists()
            if not allowed:
                raise PermissionDenied("Not allowed for this course.")
            teacher = serializer.validated_data.get("teacher")
            if teacher and teacher.company != user.company:
                raise PermissionDenied("Teacher must belong to the same company.")
            if teacher and not teacher.teaching_courses.filter(id=course.id).exists():
                raise PermissionDenied("Teacher is not assigned to this course.")
            auditorium = serializer.validated_data.get("auditorium")
            if auditorium and auditorium.company != user.company:
                raise PermissionDenied("Auditorium must belong to the same company.")
            student_ids = serializer.validated_data.get("student_ids", [])
            for student in student_ids:
                if student.company != user.company:
                    raise PermissionDenied("Student must belong to the same company.")
        
        # Вычисляем lessons_count из lessons_per_month * total_months
        lessons_per_month = serializer.validated_data.get("lessons_per_month")
        total_months = serializer.validated_data.get("total_months")
        if lessons_per_month and total_months:
            total_lessons = lessons_per_month * total_months
            serializer.validated_data["lessons_count"] = total_lessons
        
        # Вычисляем end_date с учётом нового lessons_count
        end_date = compute_group_end_date(
            serializer.validated_data.get("start_date"),
            serializer.validated_data.get("schedule_days", ""),
            serializer.validated_data.get("lessons_count"),
        )
        
        # Проверяем, что дни заполнены
        schedule_days = serializer.validated_data.get("schedule_days", "")
        if not schedule_days:
            raise PermissionDenied("Укажите дни занятий.")
        
        # Проверяем доступность аудитории
        serializer.validated_data["end_date"] = end_date
        self._ensure_auditorium_available(serializer)
        # Проверяем доступность преподавателя
        self._ensure_teacher_available(serializer)
        
        save_kwargs = {"company": user.company, "end_date": end_date}
        # If teacher is assigned, set status to pending for teacher confirmation
        teacher = serializer.validated_data.get("teacher")
        if teacher:
            save_kwargs["status"] = Group.Status.PENDING
        
        # Получаем teacher_percent из данных формы (по умолчанию 0)
        teacher_percent = serializer.validated_data.get("teacher_percent", 0) or 0
        save_kwargs["teacher_percent"] = teacher_percent
        
        serializer.save(**save_kwargs)
        
        # Автоматическое начисление % учителю при создании группы
        # Формула: кол-во студентов × цена курса × teacher_percent / 100
        if teacher and course and teacher_percent and teacher_percent > 0:
            try:
                teacher_user = User.objects.get(id=teacher.id)
                course_price = course.price or 0
                student_count = serializer.instance.students.count() or 0
                total_amount = course_price * student_count
                percent_amount = (total_amount * teacher_percent) / 100

                if percent_amount > 0:
                    teacher_balance, created = UserBalance.objects.get_or_create(user=teacher_user)
                    reason = (
                        f"Начисление %{teacher_percent}% от группы «{serializer.instance.name}» "
                        f"({student_count} студ. × {course_price} eC)"
                    )
                    teacher_balance.add_coins(int(percent_amount), reason)
                    logger.info(
                        f"Teacher {teacher_user.username} credited {int(percent_amount)} eC "
                        f"({teacher_percent}% of {student_count} students × {course_price} eC = {total_amount} eC)"
                    )
            except Exception as e:
                logger.warning(f"Failed to credit teacher percent on create: {e}")
        
        # Авто-создание GroupMonth (месяцев обучения) при создании группы
        months_count = total_months or 0
        self._create_group_months(serializer.instance, months_count)
        
        # Авто-создание договоров для студентов группы
        if course:
            student_ids_list = serializer.validated_data.get("student_ids", [])
            if student_ids_list:
                for student in student_ids_list:
                    # Проверяем, что договор ещё не существует
                    if not Contract.objects.filter(student=student, group=serializer.instance).exists():
                        try:
                            Contract.objects.create(
                                company=user.company,
                                student=student,
                                group=serializer.instance,
                                amount=course.price or 0,
                                start_date=serializer.instance.start_date or timezone.now().date(),
                                end_date=serializer.instance.end_date,
                                created_by=user,
                                status=Contract.Status.DRAFT,
                            )
                            logger.info(f"Auto-contract created for student {student.id} in group {serializer.instance.name}")
                        except Exception as e:
                            logger.warning(f"Failed to auto-create contract for student {student.id}: {e}")
        
        # Отправить уведомление учителю, если группа назначена
        if teacher:
            try:
                from telegram_bot.notifications import send_group_request_notification
                from asgiref.sync import async_to_sync
                
                # Получаем свежий объект группы
                group = Group.objects.select_related('teacher').get(id=serializer.instance.id)
                if group.teacher and group.teacher.telegram_chat_id:
                    async_to_sync(send_group_request_notification)(group)
            except Exception as e:
                logger.warning(f"Failed to send group notification to teacher: {e}")

    def _create_group_months(self, group: Group, months_count: int):
        """
        Синхронизирует месяцы обучения группы: создаёт недостающие,
        удаляет лишние (если total_months уменьшился).
        """
        if months_count < 1:
            # Если total_months = 0, удаляем все месяцы группы
            GroupMonth.objects.filter(group=group).delete()
            return

        # Удаляем лишние месяцы (если total_months уменьшился)
        GroupMonth.objects.filter(
            group=group,
            month_number__gt=months_count,
        ).delete()

        # Создаём недостающие месяцы
        for month_number in range(1, months_count + 1):
            GroupMonth.objects.get_or_create(
                group=group,
                month_number=month_number,
                defaults={"status": GroupMonth.Status.PENDING},
            )

    def perform_update(self, serializer):
        user = self.request.user
        instance = self.get_object()
        teacher_changed = False

        # Сохраняем старые значения до обновления
        old_teacher_percent = instance.teacher_percent
        old_course = instance.course

        if user.role in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            if user.role == User.Role.MANAGER and "is_login_allowed" in serializer.validated_data:
                raise PermissionDenied("Managers cannot change group login access.")
            course = serializer.validated_data.get("course", None)
            if course:
                allowed = course.admins.filter(id=user.id).exists()
                if user.role == User.Role.MANAGER:
                    allowed = course.admins.filter(
                        company=user.company
                    ).exists()
                if not allowed:
                    raise PermissionDenied("Not allowed for this course.")
            selected_teacher = serializer.validated_data.get("teacher", instance.teacher)
            if selected_teacher and selected_teacher.company != user.company:
                raise PermissionDenied("Teacher must belong to the same company.")
            course_for_teacher = course or instance.course
            if (
                selected_teacher
                and course_for_teacher
                and not selected_teacher.teaching_courses.filter(
                    id=course_for_teacher.id
                ).exists()
            ):
                raise PermissionDenied("Teacher is not assigned to this course.")
            auditorium = serializer.validated_data.get("auditorium")
            if auditorium and auditorium.company != user.company:
                raise PermissionDenied("Auditorium must belong to the same company.")
            student_ids = serializer.validated_data.get("student_ids", [])
            for student in student_ids:
                if student.company != user.company:
                    raise PermissionDenied("Student must belong to the same company.")
            # Check if teacher changed
            if "teacher" in serializer.validated_data and serializer.validated_data["teacher"] != instance.teacher:
                teacher_changed = True
        self._ensure_auditorium_available(serializer, instance=instance)
        self._ensure_teacher_available(serializer, instance=instance)
        start_date = serializer.validated_data.get("start_date", instance.start_date)
        schedule_days = serializer.validated_data.get(
            "schedule_days", instance.schedule_days
        )
        
        # Вычисляем lessons_count из lessons_per_month * total_months (при создании или обновлении)
        lessons_per_month = serializer.validated_data.get("lessons_per_month")
        total_months = serializer.validated_data.get("total_months")
        if lessons_per_month and total_months:
            total_lessons = lessons_per_month * total_months
            serializer.validated_data["lessons_count"] = total_lessons
            lessons_count = total_lessons
        else:
            lessons_count = serializer.validated_data.get(
                "lessons_count", instance.lessons_count
            )
        
        end_date = compute_group_end_date(start_date, schedule_days, lessons_count)
        save_kwargs = {"end_date": end_date}
        # If teacher changed, reset to pending for new teacher confirmation
        if teacher_changed and instance.status != Group.Status.PENDING:
            save_kwargs["status"] = Group.Status.PENDING
        serializer.save(**save_kwargs)
        
        # Создаём месяцы обучения, если total_months или lessons_per_month были изменены
        month_fields_changed = "total_months" in serializer.validated_data or "lessons_per_month" in serializer.validated_data
        if month_fields_changed:
            final_total_months = total_months or instance.total_months
            self._create_group_months(instance, final_total_months or 0)

        # Начисляем % учителю, если изменился teacher, course или teacher_percent
        percent_fields_changed = (
            "teacher" in serializer.validated_data
            or "course" in serializer.validated_data
            or "teacher_percent" in serializer.validated_data
        )
        if percent_fields_changed:
            self._credit_teacher_percent(
                instance,
                teacher_changed=teacher_changed,
                old_percent=old_teacher_percent,
                old_course=old_course,
            )

    def _credit_teacher_percent(
        self,
        group: Group,
        teacher_changed: bool = False,
        old_percent: float = 0,
        old_course=None,
    ):
        """
        Зачисляет учителю % от общей стоимости группы.
        Формула: students_count × course_price × teacher_percent / 100.
        При teacher_changed — сначала списываем старую сумму, потом начисляем новую.
        """
        teacher = group.teacher
        course = group.course
        teacher_percent = group.teacher_percent or 0

        if not teacher or not course or not teacher_percent or teacher_percent <= 0:
            return

        student_count = group.students.count() or 0
        if student_count == 0:
            return

        try:
            teacher_user = User.objects.get(id=teacher.id)
            course_price = course.price or 0
            total_amount = course_price * student_count
            percent_amount = int((total_amount * teacher_percent) / 100)

            if percent_amount <= 0:
                return

            teacher_balance, created = UserBalance.objects.get_or_create(user=teacher_user)

            if (teacher_changed or old_percent != teacher_percent or old_course != course) and old_percent and old_percent > 0:
                # Списываем старую сумму
                old_total = (old_course.price or 0) * student_count
                old_amount = int((old_total * old_percent) / 100)
                if old_amount > 0:
                    teacher_balance.spend_coins(old_amount, f"Корректировка: списание %{old_percent}% группы «{group.name}»")
                    logger.info(f"Teacher {teacher_user.username} debited {old_amount} eC (correction for group {group.name})")

            # Начисляем новую сумму
            reason = (
                f"Начисление %{teacher_percent}% от группы «{group.name}» "
                f"({student_count} студ. × {course_price} eC)"
            )
            teacher_balance.add_coins(percent_amount, reason)
            logger.info(
                f"Teacher {teacher_user.username} credited {percent_amount} eC "
                f"({teacher_percent}% of {student_count} students × {course_price} eC = {total_amount} eC)"
            )
        except Exception as e:
            logger.warning(f"Failed to credit teacher percent for group {group.id}: {e}")

    def destroy(self, request, *args, **kwargs):
        if request.user.role == User.Role.MANAGER:
            raise PermissionDenied("Managers cannot delete groups.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["post"])
    def resubmit_after_rejection(self, request, pk=None):
        """Повторная отправка запроса учителю после отказа (только для course_admin)"""
        user = request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            raise PermissionDenied("Только курс-админ или менеджер может повторить отправку.")

        group = self.get_object()
        
        if group.status != Group.Status.REJECTED:
            return Response(
                {"detail": "Можно повторить отправку только для отклонённых групп."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        if not group.teacher:
            return Response(
                {"detail": "У группы должен быть назначен учитель."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        group.status = Group.Status.PENDING
        group.save()
        
        # Отправить уведомление учителю через Telegram
        from telegram_bot.helpers import _get_group_by_id
        from telegram_bot.notifications import send_group_request_notification
        from asgiref.sync import async_to_sync
        
        try:
            group_with_teacher = Group.objects.select_related('teacher').get(id=group.id)
            async_to_sync(send_group_request_notification)(group_with_teacher)
        except Exception as e:
            logger.warning(f"Failed to send group request notification: {e}")

        return Response({
            "detail": "Запрос повторно отправлен учителю.",
            "status": group.status
        })

    def _ensure_auditorium_available(self, serializer, instance=None):
        auditorium = serializer.validated_data.get(
            "auditorium",
            instance.auditorium if instance else None,
        )
        schedule_time = serializer.validated_data.get(
            "schedule_time",
            instance.schedule_time if instance else "",
        )
        schedule_days = serializer.validated_data.get(
            "schedule_days",
            instance.schedule_days if instance else "",
        )
        start_date = serializer.validated_data.get(
            "start_date",
            instance.start_date if instance else None,
        )
        end_date = serializer.validated_data.get(
            "end_date",
            instance.end_date if instance else None,
        )
        course = serializer.validated_data.get(
            "course",
            instance.course if instance else None,
        )
        if not auditorium or not schedule_time or not schedule_days or not course:
            return
        duration = course.lesson_duration_minutes or None
        if not duration:
            return
        start_minutes = parse_time_to_minutes(schedule_time)
        if start_minutes is None:
            return
        end_minutes = start_minutes + duration
        days_set = parse_schedule_days(schedule_days)
        if not days_set:
            return

        qs = Group.objects.filter(auditorium=auditorium)
        if instance:
            qs = qs.exclude(id=instance.id)

        for group in qs:
            if not group.schedule_time or not group.schedule_days:
                continue
            other_duration = (
                group.course.lesson_duration_minutes if group.course else None
            )
            if not other_duration:
                continue
            other_start = parse_time_to_minutes(group.schedule_time)
            if other_start is None:
                continue
            other_end = other_start + other_duration
            if not time_ranges_overlap(
                start_minutes, end_minutes, other_start, other_end
            ):
                continue
            other_days = parse_schedule_days(group.schedule_days)
            if not other_days or not days_set.intersection(other_days):
                continue
            if not ranges_overlap(
                start_date, end_date, group.start_date, group.end_date
            ):
                continue
            raise PermissionDenied("Auditorium is busy at this time.")

    def _ensure_teacher_available(self, serializer, instance=None):
        """
        Проверяет, что преподаватель не занят в другой группе
        в то же время, дни недели и даты.
        """
        teacher = serializer.validated_data.get(
            "teacher",
            instance.teacher if instance else None,
        )
        if not teacher:
            return
        schedule_time = serializer.validated_data.get(
            "schedule_time",
            instance.schedule_time if instance else "",
        )
        schedule_days = serializer.validated_data.get(
            "schedule_days",
            instance.schedule_days if instance else "",
        )
        start_date = serializer.validated_data.get(
            "start_date",
            instance.start_date if instance else None,
        )
        end_date = serializer.validated_data.get(
            "end_date",
            instance.end_date if instance else None,
        )
        course = serializer.validated_data.get(
            "course",
            instance.course if instance else None,
        )
        if not schedule_time or not schedule_days or not course:
            return
        duration = course.lesson_duration_minutes or None
        if not duration:
            return
        start_minutes = parse_time_to_minutes(schedule_time)
        if start_minutes is None:
            return
        end_minutes = start_minutes + duration
        days_set = parse_schedule_days(schedule_days)
        if not days_set:
            return

        qs = Group.objects.filter(teacher=teacher)
        if instance:
            qs = qs.exclude(id=instance.id)

        for group in qs:
            if not group.schedule_time or not group.schedule_days:
                continue
            other_duration = (
                group.course.lesson_duration_minutes if group.course else None
            )
            if not other_duration:
                continue
            other_start = parse_time_to_minutes(group.schedule_time)
            if other_start is None:
                continue
            other_end = other_start + other_duration
            if not time_ranges_overlap(
                start_minutes, end_minutes, other_start, other_end
            ):
                continue
            other_days = parse_schedule_days(group.schedule_days)
            if not other_days or not days_set.intersection(other_days):
                continue
            if not ranges_overlap(
                start_date, end_date, group.start_date, group.end_date
            ):
                continue
            raise PermissionDenied(
                f"Преподаватель «{teacher}» уже занят в это время в группе «{group.name}»."
            )


class AuditoriumViewSet(viewsets.ModelViewSet):
    queryset = Auditorium.objects.all().order_by("-created_at")
    serializer_class = AuditoriumSerializer
    permission_classes = [IsCourseAdminOrManagerReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == User.Role.COURSE_ADMIN:
            if user.company:
                return queryset.filter(company=user.company)
            return queryset.none()
        if user.is_authenticated and user.role == User.Role.MANAGER:
            if user.company:
                return queryset.filter(company=user.company)
            return queryset.none()
        return queryset.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == User.Role.MANAGER:
            raise PermissionDenied("Managers cannot create auditoriums.")
        company = user.company
        serializer.save(company=company)


class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all().order_by("-created_at")
    serializer_class = AttendanceSerializer
    permission_classes = [IsCourseAdminOrTeacherReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == User.Role.COURSE_ADMIN:
            return queryset.filter(
                models.Q(group__course__admins=user)
                | models.Q(group__company=user.company)
            ).distinct()
        if user.is_authenticated and user.role == User.Role.TEACHER:
            return queryset.filter(group__teacher=user)
        if user.is_authenticated and user.role == User.Role.STUDENT:
            return queryset.filter(student__user=user)
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        if user.role in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            raise permissions.PermissionDenied("Course admins cannot mark attendance.")
        group = serializer.validated_data.get("group")
        if user.role == User.Role.TEACHER and group.teacher_id != user.id:
            raise permissions.PermissionDenied("Not allowed for this group.")
        serializer.save()


class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.all().order_by("-created_at")
    serializer_class = PaymentSerializer
    permission_classes = [IsCourseAdminOrManagerOrStudentReadOnly]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == User.Role.COURSE_ADMIN:
            return queryset.filter(company__owner=user).distinct()
        if user.is_authenticated and user.role == User.Role.MANAGER:
            if not user.company:
                return queryset.none()
            return queryset.filter(company=user.company).distinct()
        if user.is_authenticated and user.role == User.Role.STUDENT:
            return queryset.filter(student__user=user)
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        if user.role in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            student = serializer.validated_data.get("student")
            group = serializer.validated_data.get("group")
            allowed = False
            if student and student.primary_course:
                if user.role == User.Role.COURSE_ADMIN:
                    if student.primary_course.admins.filter(id=user.id).exists():
                        allowed = True
                else:
                    if student.primary_course.admins.filter(
                        company=user.company
                    ).exists():
                        allowed = True
            if student and student.company == user.company:
                allowed = True
            if group:
                if group.course:
                    if user.role == User.Role.COURSE_ADMIN:
                        if group.course.admins.filter(id=user.id).exists():
                            allowed = True
                    else:
                        if group.course.admins.filter(
                            company=user.company
                        ).exists():
                            allowed = True
                if group.company and group.company == user.company:
                    allowed = True
            if not allowed:
                raise PermissionDenied("Not allowed for this course.")
        elif user.role == User.Role.STUDENT:
            raise PermissionDenied("Students cannot create payments.")
        student = serializer.validated_data.get("student")
        group = serializer.validated_data.get("group")
        company = serializer.validated_data.get("company")
        if not company:
            company = (
                (student.company if student and student.company else None)
                or (group.company if group and group.company else None)
            )
        serializer.save(company=company)

    def destroy(self, request, *args, **kwargs):
        if request.user.role in (
            User.Role.COURSE_ADMIN,
            User.Role.MANAGER,
            User.Role.STUDENT,
        ):
            raise PermissionDenied("Not allowed to delete payments.")
        return super().destroy(request, *args, **kwargs)


class DashboardView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        total_students = Student.objects.count()
        total_income = (
            Payment.objects.filter(status=Payment.Status.PAID).aggregate(
                total=Sum("amount")
            )["total"]
            or 0
        )
        total_debt = (
            Payment.objects.filter(status=Payment.Status.DEBT).aggregate(
                total=Sum("amount")
            )["total"]
            or 0
        )
        return Response(
            {
                "total_students": total_students,
                "total_income": total_income,
                "total_debt": total_debt,
            }
        )


class SuperAdminStatsView(APIView):
    """Полная статистика для супер-админа"""
    permission_classes = [IsAdmin]

    def get(self, request):
        # Фильтры
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        company_id = request.query_params.get('company_id')
        search = request.query_params.get('search', '').strip()
        
        # Основная статистика
        companies = Company.objects.all()
        
        # Фильтр компаний
        if company_id:
            companies = companies.filter(id=company_id)
        if search:
            companies = companies.filter(name__icontains=search)
        
        companies_count = companies.count()
        
        students = Student.objects.all()
        students_count = students.count()
        
        users = User.objects.all()
        superadmins = users.filter(role=User.Role.SUPER_ADMIN).count()
        admins = users.filter(role=User.Role.ADMIN).count()
        course_admins = users.filter(role=User.Role.COURSE_ADMIN).count()
        managers = users.filter(role=User.Role.MANAGER).count()
        teachers = users.filter(role=User.Role.TEACHER).count()
        students_users = users.filter(role=User.Role.STUDENT).count()
        
        # Балансы
        balances = CompanyBalance.objects.all()
        total_balance = sum(b.balance for b in balances)
        
        # Средний баланс
        avg_balance = total_balance / companies.count() if companies.count() > 0 else 0
        
        # Транзакции
        transactions_count = Transaction.objects.count()
        
        # Группы
        groups_count = Group.objects.count()
        
        # Курсы
        courses_count = Course.objects.count()
        
        # Публичные курсы
        public_courses_count = PublicCourse.objects.all()
        if date_from or date_to:
            from django.utils import timezone
            from datetime import datetime
            if date_from:
                public_courses_count = public_courses_count.filter(created_at__gte=date_from)
            if date_to:
                public_courses_count = public_courses_count.filter(created_at__lte=date_to)
        public_courses_count = public_courses_count.count()
        
        # Аудитории
        auditoriums_count = Auditorium.objects.count()
        
        # Платежи с фильтрами по дате
        payments = Payment.objects.all()
        if date_from:
            payments = payments.filter(paid_at__gte=date_from)
        if date_to:
            payments = payments.filter(paid_at__lte=date_to)
        payments_count = payments.count()
        paid_count = payments.filter(status=Payment.Status.PAID).count()
        debt_count = payments.filter(status=Payment.Status.DEBT).count()
        total_paid = payments.filter(status=Payment.Status.PAID).aggregate(Sum('amount'))['amount__sum'] or 0
        
        # Task Leads
        task_leads_count = TaskLead.objects.count()
        
        # Задачи с фильтрами
        tasks = Task.objects.all()
        if date_from:
            tasks = tasks.filter(created_at__gte=date_from)
        if date_to:
            tasks = tasks.filter(created_at__lte=date_to)
        tasks_count = tasks.count()
        pending_tasks = tasks.filter(status=Task.Status.PENDING).count()
        in_progress_tasks = tasks.filter(status=Task.Status.IN_PROGRESS).count()
        completed_tasks = tasks.filter(status=Task.Status.COMPLETED).count()
        
        # Домашние задания
        homework_tasks_count = HomeworkTask.objects.count()
        
        # Сданные задания
        submissions = HomeworkSubmission.objects.all()
        submissions_count = submissions.count()
        reviewed_submissions = submissions.filter(status=HomeworkSubmission.Status.REVIEWED).count()
        pending_submissions = submissions.filter(status=HomeworkSubmission.Status.PENDING).count()
        
        # Trial Leads
        trial_leads = TrialLead.objects.all()
        if date_from:
            trial_leads = trial_leads.filter(created_at__gte=date_from)
        if date_to:
            trial_leads = trial_leads.filter(created_at__lte=date_to)
        trial_leads_count = trial_leads.count()
        converted_trial_leads = trial_leads.filter(status=TrialLead.Status.CONVERTED).count()
        
        # Конверсия пробных уроков
        conversion_rate = (converted_trial_leads / trial_leads_count * 100) if trial_leads_count > 0 else 0
        
        # Вакансии
        vacancies_count = JobVacancy.objects.all().count()
        
        # Новые регистрации за сегодня
        today = timezone.now().date()
        new_students_today = Student.objects.filter(created_at__date=today).count()
        new_companies_today = Company.objects.filter(created_at__date=today).count()
        new_users_today = User.objects.filter(date_joined__date=today).count()
        
        # Компании без баланса
        companies_no_balance = companies.filter(id__in=CompanyBalance.objects.values('company').annotate(count=models.Count('id')).filter(count=0).values('company')).count()
        
        # Детализация по компаниям
        companies_data = []
        for company in companies:
            balance = CompanyBalance.objects.filter(company=company).first()
            bal = balance.balance if balance else 0
            students_count_company = company.students.count()
            managers_count_company = company.users.filter(role=User.Role.MANAGER).count()
            course_admins_company = company.users.filter(role=User.Role.COURSE_ADMIN).count()
            teachers_count_company = company.users.filter(role=User.Role.TEACHER).count()
            groups_count_company = company.groups.count()
            is_active = company.is_active
            
            companies_data.append({
                "id": company.id,
                "name": company.name,
                "slug": company.slug,
                "city": company.city,
                "category": company.category,
                "students_count": students_count_company,
                "managers_count": managers_count_company,
                "course_admins_count": course_admins_company,
                "teachers_count": teachers_count_company,
                "groups_count": groups_count_company,
                "balance": bal,
                "is_active": is_active,
                "created_at": company.created_at.isoformat(),
            })
        
        # Данные для графиков (динамика по месяцам)
        from django.db.models import Count, Q
        from django.db.models.functions import TruncMonth
        
        # Студенты по месяцам
        monthly_students = Student.objects.annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')
        
        monthly_students_data = [{'month': m['month'].isoformat(), 'count': m['count']} for m in monthly_students[:12]]
        
        # Платежи по месяцам
        monthly_payments = Payment.objects.filter(status=Payment.Status.PAID).annotate(
            month=TruncMonth('paid_at')
        ).values('month').annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('month')
        
        monthly_payments_data = [{'month': m['month'].isoformat(), 'total': m['total'] or 0, 'count': m['count']} for m in monthly_payments[:12]]
        
        return Response({
            # Основная статистика
            "total_companies": companies_count,
            "total_students": students_count,
            "total_users": users.count(),
            "total_balance": total_balance,
            "avg_balance": round(avg_balance, 2),
            
            # Пользователи по ролям
            "superadmins": superadmins,
            "admins": admins,
            "course_admins": course_admins,
            "managers": managers,
            "teachers": teachers,
            "students_users": students_users,
            
            # Новые регистрации за сегодня
            "new_students_today": new_students_today,
            "new_companies_today": new_companies_today,
            "new_users_today": new_users_today,
            
            # Другие данные
            "transactions": transactions_count,
            "groups": groups_count,
            "courses": courses_count,
            "public_courses": public_courses_count,
            "auditoriums": auditoriums_count,
            
            # Платежи
            "payments_total": payments_count,
            "payments_paid": paid_count,
            "payments_debt": debt_count,
            "total_paid": total_paid,
            
            # Задачи
            "task_leads": task_leads_count,
            "tasks_total": tasks_count,
            "tasks_pending": pending_tasks,
            "tasks_in_progress": in_progress_tasks,
            "tasks_completed": completed_tasks,
            
            # Домашние задания
            "homework_tasks": homework_tasks_count,
            "submissions_total": submissions_count,
            "submissions_reviewed": reviewed_submissions,
            "submissions_pending": pending_submissions,
            
            # Trial Leads
            "trial_leads_total": trial_leads_count,
            "trial_leads_converted": converted_trial_leads,
            "conversion_rate": round(conversion_rate, 2),
            
            # Вакансии
            "vacancies": vacancies_count,
            
            # Статус компаний
            "companies_no_balance": companies_no_balance,
            
            # Детализация по компаниям
            "companies": companies_data,
            
            # Графики
            "charts": {
                "monthly_students": monthly_students_data,
                "monthly_payments": monthly_payments_data,
            }
        })


class LandingPageViewSet(viewsets.ModelViewSet):
    queryset = LandingPage.objects.all().prefetch_related("sections")
    serializer_class = LandingPageSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.role == User.Role.COURSE_ADMIN:
            if user.company:
                return queryset.filter(company=user.company)
            return queryset.filter(company=user.company)
        if user.role == User.Role.ADMIN or user.is_superuser:
            status_filter = self.request.query_params.get("status", "").strip()
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            company_name = self.request.query_params.get("company_name", "").strip()
            if company_name:
                queryset = queryset.filter(company=company_name)
            return queryset
        return queryset.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role != User.Role.COURSE_ADMIN:
            raise PermissionDenied("Only course admins can create landing pages.")
        if not user.can_create_landing_page():
            raise PermissionDenied(
                f"Landing pages limit reached. Maximum: {user.max_pages}, "
                f"Current: {user.get_pages_count()}"
            )
        serializer.save(owner=user, company=user.company)

    def perform_update(self, serializer):
        page = self.get_object()
        user = self.request.user
        if user.role == User.Role.COURSE_ADMIN:
            if page.company != user.company:
                raise PermissionDenied("Not allowed for this landing page.")
            serializer.save()
            return
        if user.role == User.Role.ADMIN or user.is_superuser:
            serializer.save()
            return
        raise PermissionDenied("Not allowed.")

    def destroy(self, request, *args, **kwargs):
        page = self.get_object()
        user = request.user
        if user.role == User.Role.COURSE_ADMIN and page.company != user.company:
            raise PermissionDenied("Not allowed for this landing page.")
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.ADMIN) and not user.is_superuser:
            raise PermissionDenied("Not allowed to delete this landing page.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["post"], url_path="submit")
    def submit(self, request, pk=None):
        page = self.get_object()
        user = request.user
        if user.role != User.Role.COURSE_ADMIN or page.company != user.company:
            raise PermissionDenied("Only the owning course admin can submit this landing page.")
        if page.status == LandingPage.Status.PENDING:
            raise PermissionDenied("This landing page is already pending moderation.")
        validate_landing_page_for_publication(page, user)
        page.status = LandingPage.Status.PENDING
        page.moderation_comment = ""
        page.submitted_at = timezone.now()
        page.save(update_fields=["status", "moderation_comment", "submitted_at", "updated_at"])
        return Response(self.get_serializer(page).data)

    @action(detail=True, methods=["post"], url_path="approve")
    def approve(self, request, pk=None):
        page = self.get_object()
        user = request.user
        if user.role != User.Role.ADMIN and not user.is_superuser:
            raise PermissionDenied("Only admins can approve landing pages.")
        if page.status != LandingPage.Status.PENDING:
            raise PermissionDenied("Only pending landing pages can be approved.")
        validate_landing_page_for_publication(page, page.owner)
        page.status = LandingPage.Status.ACTIVE
        page.moderation_comment = ""
        page.moderated_at = timezone.now()
        page.moderated_by = user
        page.published_at = page.published_at or timezone.now()
        page.save(
            update_fields=[
                "status",
                "moderation_comment",
                "moderated_at",
                "moderated_by",
                "published_at",
                "updated_at",
            ]
        )
        return Response(self.get_serializer(page).data)

    @action(detail=True, methods=["post"], url_path="reject")
    def reject(self, request, pk=None):
        page = self.get_object()
        user = request.user
        if user.role != User.Role.ADMIN and not user.is_superuser:
            raise PermissionDenied("Only admins can reject landing pages.")
        if page.status != LandingPage.Status.PENDING:
            raise PermissionDenied("Only pending landing pages can be rejected.")
        comment = (request.data.get("comment") or "").strip()
        if not comment:
            return Response(
                {"detail": "Moderation comment is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        page.status = LandingPage.Status.REJECTED
        page.moderation_comment = comment
        page.moderated_at = timezone.now()
        page.moderated_by = user
        page.save(
            update_fields=[
                "status",
                "moderation_comment",
                "moderated_at",
                "moderated_by",
                "updated_at",
            ]
        )
        return Response(self.get_serializer(page).data)


class LandingHeaderLinkViewSet(viewsets.ModelViewSet):
    queryset = LandingHeaderLink.objects.all().select_related("target_page")
    serializer_class = LandingHeaderLinkSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.role == User.Role.COURSE_ADMIN:
            return queryset.filter(company=user.company)
        if user.role == User.Role.ADMIN or user.is_superuser:
            company_name = self.request.query_params.get("company_name", "").strip()
            if company_name:
                queryset = queryset.filter(company=company_name)
            return queryset
        return queryset.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role != User.Role.COURSE_ADMIN:
            raise PermissionDenied("Only course admins can manage landing header links.")
        serializer.save(company=user.company)

    def perform_update(self, serializer):
        link = self.get_object()
        user = self.request.user
        if user.role != User.Role.COURSE_ADMIN or link.company != user.company:
            raise PermissionDenied("Only the owning course admin can update this header link.")
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        link = self.get_object()
        if request.user.role != User.Role.COURSE_ADMIN or link.company != request.user.company:
            raise PermissionDenied("Only the owning course admin can delete this header link.")
        return super().destroy(request, *args, **kwargs)


class PublicLandingDetailView(APIView):
    permission_classes = [permissions.AllowAny]
    throttle_classes = [AnonRateThrottle, PublicReadThrottle]

    def get(self, request, slug: str):
        page = get_object_or_404(
            LandingPage.objects.prefetch_related("sections"),
            slug=slug,
            status=LandingPage.Status.ACTIVE,
        )
        return Response(LandingPublicPageSerializer(page, context={"request": request}).data)


class PublicLandingLeadCreateView(APIView):
    permission_classes = [permissions.AllowAny]
    throttle_classes = [AnonRateThrottle, PublicSubmitThrottle]
    parser_classes = [JSONParser]

    def post(self, request, slug: str):
        page = get_object_or_404(
            LandingPage,
            slug=slug,
            status=LandingPage.Status.ACTIVE,
        )
        full_name = (request.data.get("full_name") or "").strip()
        phone = (request.data.get("phone") or "").strip()
        if not full_name or not phone:
            return Response(
                {"detail": "Full name and phone are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # Санитизация полей от HTML перед сохранением
        safe_full_name = bleach.clean(full_name, tags=[], strip=True)[:200]
        safe_course_interest = bleach.clean((request.data.get("course_interest") or "").strip(), tags=[], strip=True)[:200]
        safe_comment = bleach.clean((request.data.get("comment") or "").strip(), tags=[], strip=True)[:1000]

        lead = TrialLead.objects.create(
            full_name=safe_full_name,
            phone=phone,
            course_interest=safe_course_interest,
            source=f"landing:{page.slug}",
            comment=safe_comment,
            company=page.company,
        )
        return Response(TrialLeadSerializer(lead).data, status=status.HTTP_201_CREATED)


class CrmContactView(APIView):
    """Public endpoint for the CRM's own landing page contact form.

    Creates a TrialLead without company association and notifies superadmins.
    """
    permission_classes = [permissions.AllowAny]
    throttle_classes = [AnonRateThrottle, PublicSubmitThrottle]
    parser_classes = [JSONParser]

    def post(self, request):
        full_name = (request.data.get("full_name") or "").strip()
        phone = (request.data.get("phone") or "").strip()
        if not full_name or not phone:
            return Response(
                {"detail": "Full name and phone are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        comment = (request.data.get("comment") or "").strip()
        telegram = (request.data.get("telegram") or "").strip()

        # Санитизация полей от HTML перед сохранением
        safe_full_name = bleach.clean(full_name, tags=[], strip=True)[:200]
        safe_comment = bleach.clean(comment, tags=[], strip=True)[:1000]
        safe_telegram = bleach.clean(telegram, tags=[], strip=True)[:200]

        # Save as TrialLead with source "crm-landing" (no company)
        comment_parts = []
        if safe_comment:
            comment_parts.append(safe_comment)
        if safe_telegram:
            comment_parts.append(f"Telegram: {safe_telegram}")
        lead = TrialLead.objects.create(
            full_name=safe_full_name,
            phone=phone,
            source="crm-landing",
            comment="\n".join(comment_parts),
            company=None,
        )

        # Notify superadmins (используем санитизированные данные)
        try:
            from telegram_bot.notifications import send_crm_contact_notification
            from asgiref.sync import async_to_sync

            async_to_sync(send_crm_contact_notification)(
                full_name=safe_full_name,
                phone=phone,
                comment=safe_comment,
                telegram=safe_telegram,
            )
        except Exception as e:
            logger.warning(f"Failed to send CRM contact notification: {e}")

        return Response(
            {"id": lead.id, "detail": "Contact request received."},
            status=status.HTTP_201_CREATED,
        )


class TrialLeadViewSet(viewsets.ModelViewSet):
    queryset = TrialLead.objects.all().select_related("assignment__manager").order_by("-created_at")
    serializer_class = TrialLeadSerializer
    permission_classes = [IsCourseAdminOrManager]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == User.Role.COURSE_ADMIN:
            if user.company:
                return queryset.filter(company=user.company)
            return queryset.filter(company=user.company)
        if user.is_authenticated and user.role == User.Role.MANAGER:
            if user.company:
                return queryset.filter(company=user.company)
            return queryset.filter(company=user.company)
        return queryset.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role != User.Role.MANAGER:
            raise PermissionDenied("Only managers can create trial leads.")
        group = serializer.validated_data.get("group_assigned")
        if group and group.company != user.company:
            raise PermissionDenied("Not allowed for this group.")
        serializer.save(company=user.company)

    def perform_update(self, serializer):
        user = self.request.user
        if user.role != User.Role.MANAGER:
            raise PermissionDenied("Only managers can update trial leads.")
        group = serializer.validated_data.get("group_assigned")
        if group and group.company != user.company:
            raise PermissionDenied("Not allowed for this group.")
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        if request.user.role != User.Role.MANAGER:
            raise PermissionDenied("Only managers can delete trial leads.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=["get"], url_path="analytics")
    def analytics(self, request):
        months_param = request.query_params.get("months", "").strip()
        months = [m.strip() for m in months_param.split(",") if m.strip()]
        if not months:
            today = date.today()
            months = [f"{today.year:04d}-{today.month:02d}"]
        if len(months) < 1 or len(months) > 6:
            return Response(
                {"detail": "Months count must be between 1 and 6."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        parsed_months = []
        for value in months:
            parts = value.split("-")
            if len(parts) != 2:
                return Response(
                    {"detail": "Invalid month format. Use YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                year = int(parts[0])
                month = int(parts[1])
                if month < 1 or month > 12:
                    raise ValueError()
            except ValueError:
                return Response(
                    {"detail": "Invalid month format. Use YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            parsed_months.append((value, year, month))

        qs = self.get_queryset()
        monthly_data = []
        sources_by_month = []
        ages_by_month = []

        total_leads = 0
        attended_total = 0
        not_attended_total = 0
        converted_total = 0

        for value, year, month in parsed_months:
            last_day = monthrange(year, month)[1]
            start_date = date(year, month, 1)
            end_date = date(year, month, last_day)
            month_qs = qs.filter(
                created_at__date__gte=start_date,
                created_at__date__lte=end_date,
            )
            total = month_qs.count()
            attended = month_qs.filter(trial_attended=True).count()
            not_attended = month_qs.filter(
                status=TrialLead.Status.NOT_ATTENDED
            ).count()
            converted = month_qs.filter(
                models.Q(converted_to_student=True)
                | models.Q(status=TrialLead.Status.CONVERTED)
            ).count()
            rate = round((converted / total * 100) if total else 0, 2)

            total_leads += total
            attended_total += attended
            not_attended_total += not_attended
            converted_total += converted

            monthly_data.append(
                {
                    "month": value,
                    "total_leads": total,
                    "attended_trial": attended,
                    "not_attended": not_attended,
                    "converted_students": converted,
                    "conversion_rate": rate,
                }
            )

            sources_raw = (
                month_qs.values("source")
                .annotate(total=models.Count("id"))
                .order_by("-total")
            )
            sources_items = [
                {
                    "label": item["source"] or "—",
                    "total": item["total"],
                }
                for item in sources_raw
            ]
            sources_by_month.append({"month": value, "items": sources_items})

            ages_by_month.append(
                {
                    "month": value,
                    "items": compute_age_groups(month_qs),
                }
            )

        summary_rate = round(
            (converted_total / total_leads * 100) if total_leads else 0, 2
        )

        return Response(
            {
                "months": [value for value, _, _ in parsed_months],
                "summary": {
                    "total_leads": total_leads,
                    "attended_trial": attended_total,
                    "not_attended": not_attended_total,
                    "converted_students": converted_total,
                    "conversion_rate": summary_rate,
                },
                "monthly_data": monthly_data,
                "sources": sources_by_month,
                "age_groups": ages_by_month,
            }
        )


class TaskViewSet(viewsets.ModelViewSet):
    queryset = Task.objects.all().order_by("-created_at")
    serializer_class = TaskSerializer
    permission_classes = [IsCourseAdminOrManager]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == User.Role.COURSE_ADMIN:
            if not user.company:
                return queryset.none()
            return queryset.filter(company=user.company)
        if user.is_authenticated and user.role == User.Role.MANAGER:
            if not user.company:
                return queryset.none()
            return queryset.filter(assigned_to=user, company=user.company)
        return queryset.none()

    def create(self, request, *args, **kwargs):
        user = request.user
        if user.role != User.Role.COURSE_ADMIN:
            raise PermissionDenied("Only course admins can create tasks.")
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        assigned_to = serializer.validated_data.get("assigned_to")
        if not assigned_to or assigned_to.role != User.Role.MANAGER:
            raise PermissionDenied("Task must be assigned to a manager.")
        if resolve_user_company_name(assigned_to) != resolve_user_company_name(user):
            raise PermissionDenied("Manager must belong to the same company.")

        tasks = build_task_instances(serializer.validated_data, user)
        Task.objects.bulk_create(tasks)
        data = TaskSerializer(tasks, many=True).data
        return Response(data, status=status.HTTP_201_CREATED)

    def perform_update(self, serializer):
        user = self.request.user
        if user.role == User.Role.MANAGER:
            # Managers can only update status of their tasks
            if self.get_object().assigned_to_id != user.id:
                raise PermissionDenied("Not allowed for this task.")
            allowed_fields = {"status", "is_seen"}
            update_fields = set(serializer.validated_data.keys())
            if not update_fields.issubset(allowed_fields):
                raise PermissionDenied("Managers can only update status or seen flag.")
            serializer.save()
            return
        if user.role == User.Role.COURSE_ADMIN:
            assigned_to = serializer.validated_data.get("assigned_to")
            if assigned_to and resolve_user_company_name(assigned_to) != resolve_user_company_name(user):
                raise PermissionDenied("Manager must belong to the same company.")
            serializer.save()
            return
        raise PermissionDenied("Not allowed.")

    def destroy(self, request, *args, **kwargs):
        if request.user.role != User.Role.COURSE_ADMIN:
            raise PermissionDenied("Only course admins can delete tasks.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=["post"], url_path="mark-seen")
    def mark_seen(self, request):
        user = request.user
        if user.role != User.Role.MANAGER:
            raise PermissionDenied("Only managers can mark tasks as seen.")
        data = request.data
        ids = []
        if isinstance(data, dict):
            ids = data.get("ids", []) or []
        elif isinstance(data, list):
            ids = data
        elif isinstance(data, str):
            # Allow plain payloads like "1,2,3" or "5"
            raw = data.strip()
            if raw:
                if "," in raw:
                    ids = [item.strip() for item in raw.split(",") if item.strip()]
                else:
                    ids = [raw]

        normalized_ids = []
        for item in ids:
            try:
                normalized_ids.append(int(item))
            except (TypeError, ValueError):
                continue

        queryset = self.get_queryset()
        if normalized_ids:
            queryset = queryset.filter(id__in=normalized_ids)
        updated = queryset.update(is_seen=True)
        return Response({"updated": updated})


def validate_landing_page_for_publication(page: LandingPage, owner: User | None):
    if owner and owner.role == User.Role.COURSE_ADMIN and page.sections.count() > owner.max_blocks:
        raise PermissionDenied(
            f"Page exceeds the allowed number of blocks ({owner.max_blocks})."
        )
    total_pages = LandingPage.objects.filter(company=page.company).count()
    if total_pages > 1:
        links = LandingHeaderLink.objects.filter(company=page.company)
        if not links.exists():
            raise PermissionDenied(
                "Header navigation must be configured when more than one landing page exists."
            )
        invalid_target_exists = links.exclude(target_page__company=page.company).exists()
        if invalid_target_exists:
            raise PermissionDenied("All header links must target pages from the same company.")


class HomeworkTaskViewSet(viewsets.ModelViewSet):
    queryset = HomeworkTask.objects.all().select_related("group", "teacher").prefetch_related("attachments", "students", "submissions").order_by("-created_at")
    serializer_class = HomeworkTaskSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == User.Role.STUDENT:
            now = timezone.now()
            return queryset.filter(
                is_published=True,
            ).filter(
                models.Q(publish_at__isnull=True) | models.Q(publish_at__lte=now)
            ).filter(
                models.Q(
                    target_type=HomeworkTask.TargetType.ALL_GROUP,
                    group__students__user=user,
                )
                | models.Q(
                    target_type=HomeworkTask.TargetType.SPECIFIC_STUDENTS,
                    students__user=user,
                )
            ).distinct()
        if user.is_authenticated and user.role == User.Role.TEACHER:
            return queryset.filter(teacher=user)
        if user.is_authenticated and user.role == User.Role.COURSE_ADMIN:
            return queryset.filter(company=user.company)
        return queryset.none()

    def create(self, request, *args, **kwargs):
        user = request.user
        if user.role != User.Role.TEACHER:
            raise PermissionDenied("Only teachers can create homework tasks.")
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        group = serializer.validated_data.get("group")
        if not group or group.teacher_id != user.id:
            raise PermissionDenied("Homework can only be created for your own groups.")
        instance = serializer.save(teacher=user, company=user.company)
        self._save_attachments(instance)
        data = self.get_serializer(instance).data
        headers = self.get_success_headers(data)
        return Response(data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_update(self, serializer):
        user = self.request.user
        instance = self.get_object()
        if user.role == User.Role.TEACHER:
            if instance.teacher_id != user.id:
                raise PermissionDenied("Not allowed for this homework task.")
            group = serializer.validated_data.get("group", instance.group)
            if group.teacher_id != user.id:
                raise PermissionDenied("Homework can only belong to your own groups.")
            instance = serializer.save()
            self._save_attachments(instance, replace=True)
            return
        if user.role == User.Role.COURSE_ADMIN:
            if instance.company != user.company:
                raise PermissionDenied("Not allowed for this homework task.")
            updated = serializer.save()
            self._save_attachments(updated, replace=True)
            return
        raise PermissionDenied("Not allowed.")

    def destroy(self, request, *args, **kwargs):
        user = request.user
        instance = self.get_object()
        if user.role == User.Role.TEACHER and instance.teacher_id == user.id:
            return super().destroy(request, *args, **kwargs)
        if user.role == User.Role.COURSE_ADMIN and instance.company == user.company:
            return super().destroy(request, *args, **kwargs)
        raise PermissionDenied("Not allowed to delete this homework task.")

    def _save_attachments(self, instance: HomeworkTask, replace: bool = False):
        files = self.request.FILES.getlist("files")
        if replace:
            clear_files = self.request.data.get("clear_files")
            if str(clear_files).lower() in {"1", "true", "yes"}:
                instance.attachments.all().delete()
        for file_obj in files:
            HomeworkTaskAttachment.objects.create(task=instance, file=file_obj)


class HomeworkSubmissionViewSet(viewsets.ModelViewSet):
    queryset = HomeworkSubmission.objects.all().select_related("task", "student", "student__user").order_by("-submitted_at")
    serializer_class = HomeworkSubmissionSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == User.Role.STUDENT:
            return queryset.filter(student__user=user)
        if user.is_authenticated and user.role == User.Role.TEACHER:
            return queryset.filter(task__teacher=user)
        if user.is_authenticated and user.role == User.Role.COURSE_ADMIN:
            if user.company:
                return queryset.filter(task__company=user.company)
            return queryset.none()
        if user.is_authenticated and user.role == User.Role.MANAGER:
            if user.company:
                return queryset.filter(task__company=user.company)
            return queryset.none()
        return queryset.none()

    def create(self, request, *args, **kwargs):
        user = request.user
        if user.role != User.Role.STUDENT:
            raise PermissionDenied("Only students can submit homework.")
        student = getattr(user, "student_profile", None)
        if not student:
            raise PermissionDenied("Student profile not found.")
        task_id = request.data.get("task")
        task = get_object_or_404(HomeworkTask, pk=task_id)
        if not _student_can_access_homework_task(task, student):
            raise PermissionDenied("You can submit homework only for your own groups.")
        if _is_submission_locked(task):
            raise PermissionDenied("Submission deadline has passed.")
        if HomeworkSubmission.objects.filter(task=task, student=student).exists():
            raise PermissionDenied("Submission already exists for this task.")
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(student=student)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_update(self, serializer):
        user = self.request.user
        instance = self.get_object()
        if user.role == User.Role.STUDENT:
            if instance.student.user_id != user.id:
                raise PermissionDenied("Not allowed for this submission.")
            allowed_fields = {"answer_text", "file"}
            update_fields = set(serializer.validated_data.keys())
            if not update_fields.issubset(allowed_fields):
                raise PermissionDenied("Students can only update submission content.")
            if _is_submission_locked(instance.task):
                raise PermissionDenied("Submission deadline has passed.")
            serializer.save(status=HomeworkSubmission.Status.PENDING)
            return
        if user.role == User.Role.TEACHER:
            if instance.task.teacher_id != user.id:
                raise PermissionDenied("Not allowed for this submission.")
            allowed_fields = {"status", "grade", "teacher_comment"}
            update_fields = set(serializer.validated_data.keys())
            if not update_fields.issubset(allowed_fields):
                raise PermissionDenied("Teachers can only review homework submissions.")
            serializer.save()
            return
        raise PermissionDenied("Not allowed.")

    def destroy(self, request, *args, **kwargs):
        user = request.user
        instance = self.get_object()
        if user.role == User.Role.STUDENT and instance.student.user_id == user.id:
            return super().destroy(request, *args, **kwargs)
        if user.role == User.Role.TEACHER and instance.task.teacher_id == user.id:
            return super().destroy(request, *args, **kwargs)
        raise PermissionDenied("Not allowed to delete this submission.")


def build_task_instances(validated_data, user):
    repeat_type = validated_data.get("repeat_type", Task.RepeatType.NONE)
    start_date = validated_data["due_date"]
    end_date = start_date + timedelta(days=180)
    dates = []

    if repeat_type == Task.RepeatType.DAILY:
        current = start_date
        while current <= end_date:
            dates.append(current)
            current = current + timedelta(days=1)
    elif repeat_type == Task.RepeatType.WEEKLY:
        current = start_date
        while current <= end_date:
            dates.append(current)
            current = current + timedelta(days=7)
    elif repeat_type == Task.RepeatType.MONTHLY:
        current = start_date
        while current <= end_date:
            dates.append(current)
            current = add_months(current, 1)
    else:
        dates.append(start_date)

    tasks = []
    for due_date in dates:
        tasks.append(
            Task(
                title=validated_data.get("title", ""),
                description=validated_data.get("description", ""),
                assigned_to=validated_data.get("assigned_to"),
                company=user.company,
                created_by=user,
                due_date=due_date,
                due_time=validated_data.get("due_time"),
                status=validated_data.get("status", Task.Status.PENDING),
                priority=validated_data.get("priority", Task.Priority.MEDIUM),
                repeat_type=repeat_type,
            )
        )
    return tasks


def add_months(value: date, months: int) -> date:
    month = value.month - 1 + months
    year = value.year + month // 12
    month = month % 12 + 1
    day = min(value.day, monthrange(year, month)[1])
    return date(year, month, day)
def compute_age_groups(queryset):
    buckets = [
        ("<13", 0, 12),
        ("13-17", 13, 17),
        ("18-24", 18, 24),
        ("25-34", 25, 34),
        ("35+", 35, 200),
    ]
    counts = {label: 0 for label, _, _ in buckets}
    for age in queryset.values_list("age", flat=True):
        if age is None:
            continue
        for label, start, end in buckets:
            if start <= age <= end:
                counts[label] += 1
                break
    return [{"label": label, "total": total} for label, total in counts.items()]


class PromoCodeViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing promo codes (admin and course admin).
    """
    queryset = PromoCode.objects.all().order_by("-created_at")
    serializer_class = PromoCodeSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser or user.role in (User.Role.ADMIN, User.Role.SUPER_ADMIN):
            return PromoCode.objects.all().order_by("-created_at")
        if user.role == User.Role.COURSE_ADMIN:
            return PromoCode.objects.filter(created_by=user).order_by("-created_at")
        if user.role == User.Role.MANAGER:
            return PromoCode.objects.filter(created_by__role=User.Role.ADMIN).order_by("-created_at")
        return PromoCode.objects.none()

    def create(self, request, *args, **kwargs):
        print(f"🔹 PromoCodeViewSet.create - data: {request.data}")
        print(f"🔹 Content-Type: {request.content_type}")
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        user = self.request.user
        if not (user.is_superuser or user.role in (User.Role.ADMIN, User.Role.SUPER_ADMIN)):
            raise PermissionDenied("Только админ может создавать промокоды.")
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=["post"], url_path="activate")
    def activate(self, request):
        user = request.user
        if user.role not in (User.Role.ADMIN, User.Role.COURSE_ADMIN, User.Role.MANAGER):
            raise PermissionDenied("Only admins, course admins and managers can activate promo codes.")
        
        code = request.data.get("code", "").strip()
        if not code:
            return Response(
                {"detail": "Promo code is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        try:
            promo_code = PromoCode.objects.get(code=code)
        except PromoCode.DoesNotExist:
            return Response(
                {"detail": "Promo code not found."},
                status=status.HTTP_404_NOT_FOUND,
            )
        
        # Курс-админ и менеджер могут активировать только промокоды созданные супер-админом
        if user.role in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            creator = promo_code.created_by
            creator_is_admin = (
                creator
                and (creator.is_superuser or creator.role in (User.Role.ADMIN, User.Role.SUPER_ADMIN))
            )
            if not creator_is_admin:
                return Response(
                    {"detail": "Вы можете активировать только промокоды созданные супер-админом."},
                    status=status.HTTP_403_FORBIDDEN,
                )
        
        # Проверяем срок действия
        if promo_code.expiry_date and promo_code.expiry_date < timezone.now():
            return Response(
                {"detail": "Cannot activate expired promo code."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        # Проверяем, не исчерпан ли лимит активаций
        if promo_code.current_usages >= promo_code.max_usages:
            return Response(
                {"detail": "Promo code usage limit reached."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        # Используем метод модели для активации
        if not promo_code.is_active:
            promo_code.is_active = True
            promo_code.save(update_fields=["is_active"])
        
        # Получаем или создаем баланс компании
        company_name = resolve_user_company_name(user)
        if not company_name:
            return Response(
                {"detail": "Компания не найдена."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        company_balance, created = CompanyBalance.objects.get_or_create(
            company_name=company_name,
            defaults={"balance": 0}
        )
        
        # Начисляем награду
        if promo_code.reward_type == PromoCode.RewardType.COINS:
            company_balance.add_coins(promo_code.reward_value, f"Промокод: {promo_code.code}")
        else:
            # Для бонусного лимита
            Transaction.objects.create(
                company_name=company_name,
                user=user,
                amount=0,
                reason=f"Промокод (бонус лимит): {promo_code.code}",
                transaction_type=Transaction.Type.BONUS,
            )
        
        # Обновляем счётчик использований
        promo_code.current_usages += 1
        promo_code.save(update_fields=["current_usages"])
        
        return Response(PromoCodeSerializer(promo_code).data)


class AttendanceMarkView(APIView):
    permission_classes = [IsTeacherOrCourseAdminReadOnly]

    def get(self, request):
        group_id = request.query_params.get("group")
        date_str = request.query_params.get("date")
        if not group_id or not date_str:
            return Response(
                {"detail": "group and date are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            target_date = date.fromisoformat(date_str)
        except ValueError:
            return Response(
                {"detail": "Invalid date format. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        group = get_object_or_404(Group, pk=group_id)
        user = request.user
        if user.role == User.Role.MANAGER:
            raise permissions.PermissionDenied("Not allowed for managers.")
        if user.role == User.Role.STUDENT:
            raise permissions.PermissionDenied("Not allowed for students.")
        if user.role in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            allowed = False
            if group.course:
                if user.role == User.Role.COURSE_ADMIN:
                    allowed = group.course.admins.filter(id=user.id).exists()
                else:
                    allowed = group.course.admins.filter(
                        company=user.company
                    ).exists()
            if group.company and group.company == user.company:
                allowed = True
            if not allowed:
                raise permissions.PermissionDenied("Not allowed for this course.")
        if user.role == User.Role.TEACHER and group.teacher_id != user.id:
            raise permissions.PermissionDenied("Not allowed for this group.")

        students = list(group.students.all().order_by("first_name", "last_name"))
        existing = Attendance.objects.filter(group=group, date=target_date)
        status_map = {item.student_id: item.status for item in existing}

        return Response(
            {
                "group": {"id": group.id, "name": group.name},
                "date": target_date.isoformat(),
                "students": [
                    {
                        "id": student.id,
                        "first_name": student.first_name,
                        "last_name": student.last_name,
                        "status": status_map.get(student.id),
                    }
                    for student in students
                ],
            }
        )

    def post(self, request):
        if request.user.role in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            raise permissions.PermissionDenied("Course admins cannot mark attendance.")
        group_id = request.data.get("group")
        date_str = request.data.get("date")
        items = request.data.get("items", [])
        if not group_id or not date_str:
            return Response(
                {"detail": "group and date are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            target_date = date.fromisoformat(date_str)
        except ValueError:
            return Response(
                {"detail": "Invalid date format. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        group = get_object_or_404(Group, pk=group_id)
        user = request.user
        if user.role == User.Role.MANAGER:
            raise permissions.PermissionDenied("Not allowed for managers.")
        if user.role == User.Role.STUDENT:
            raise permissions.PermissionDenied("Not allowed for students.")
        if user.role == User.Role.TEACHER and group.teacher_id != user.id:
            raise permissions.PermissionDenied("Not allowed for this group.")

        students = {student.id: student for student in group.students.all()}
        updated = []

        for item in items:
            student_id = item.get("student")
            status_value = item.get("status")
            if student_id not in students:
                continue
            if status_value not in dict(Attendance.Status.choices):
                continue
            record, _ = Attendance.objects.update_or_create(
                group=group,
                student=students[student_id],
                date=target_date,
                defaults={"status": status_value},
            )
            updated.append(record)

        return Response(
            {
                "saved": len(updated),
                "date": target_date.isoformat(),
            }
        )


# === Marketplace Views ===

class MarketplaceCompanyViewSet(viewsets.ModelViewSet):
    """ViewSet for managing companies"""
    queryset = Company.objects.all()
    serializer_class = CompanySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = Company.objects.filter(is_active=True).order_by("-created_at")
        user = self.request.user
        if user.role == User.Role.COURSE_ADMIN:
            queryset = queryset.filter(owner=user)
        elif user.role == User.Role.MANAGER:
            if user.company:
                queryset = queryset.filter(owner__company=user.company)
            else:
                queryset = queryset.none()
        elif user.role in (User.Role.TEACHER, User.Role.STUDENT):
            queryset = queryset.none()
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            raise PermissionDenied("Only course admins and managers can create companies.")
        owner = user if user.role == User.Role.COURSE_ADMIN else user.created_by
        serializer.save(owner=owner)


class MarketplaceCourseViewSet(viewsets.ModelViewSet):
    """ViewSet for managing public courses"""
    queryset = PublicCourse.objects.all().order_by("-is_promoted", "-created_at").select_related("company")
    serializer_class = PublicCourseSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        queryset = PublicCourse.objects.filter(is_active=True).select_related("company").prefetch_related("company__landing_pages")
        user = self.request.user
        
        # Public access for viewing
        if not user.is_authenticated:
            return queryset
        
        if user.role == User.Role.COURSE_ADMIN:
            return queryset.filter(company__owner=user)
        elif user.role == User.Role.MANAGER:
            company_name = resolve_user_company_name(user)
            if company_name:
                return queryset.filter(company__name=company_name)
            return queryset.none()
        elif user.role == User.Role.STUDENT:
            # Students can only view active courses
            return queryset.filter(is_active=True)
        
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            raise PermissionDenied("Only course admins and managers can create courses.")
        
        owner = user if user.role == User.Role.COURSE_ADMIN else user.created_by
        company = serializer.validated_data.get("company")
        
        if not company:
            company = Company.objects.filter(owner=owner).first()
            if not company:
                raise PermissionDenied("No company found. Create a company first.")
        
        serializer.save(company=company)

    def perform_update(self, serializer):
        user = self.request.user
        course = self.get_object()
        
        if user.role == User.Role.COURSE_ADMIN and course.company.owner != user:
            raise PermissionDenied("Not allowed for this course.")
        if user.role == User.Role.MANAGER:
            if not course.company.owner__company == user.company:
                raise PermissionDenied("Not allowed for this course.")
        
        serializer.save()


class MarketplaceJobViewSet(viewsets.ModelViewSet):
    """ViewSet for managing job vacancies"""
    queryset = JobVacancy.objects.all().order_by("-is_promoted", "-created_at")
    serializer_class = JobVacancySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        queryset = JobVacancy.objects.filter(is_active=True)
        user = self.request.user
        
        if not user.is_authenticated:
            return queryset
        
        if user.role == User.Role.COURSE_ADMIN:
            return queryset.filter(company__owner=user)
        elif user.role == User.Role.MANAGER:
            if user.company:
                return queryset.filter(company__owner__company=user.company)
            return queryset.none()
        
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            raise PermissionDenied("Only course admins and managers can create jobs.")
        
        owner = user if user.role == User.Role.COURSE_ADMIN else user.created_by
        company = serializer.validated_data.get("company")
        
        if not company:
            company = Company.objects.filter(owner=owner).first()
            if not company:
                raise PermissionDenied("No company found. Create a company first.")
        
        serializer.save(company=company)


class MyCoursesView(APIView):
    """Get all courses owned by the user"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            return Response(
                {"detail": "Доступно только для course_admin и manager."},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        if not user.company:
            return Response([])
        
        if user.role == User.Role.COURSE_ADMIN:
            courses = PublicCourse.objects.filter(company=user.company, is_active=True)
        else:
            courses = PublicCourse.objects.filter(company=user.company, is_active=True)
        
        # Add view and application counts
        data = []
        for course in courses:
            course_data = PublicCourseSerializer(course).data
            course_data['views'] = course.views
            course_data['applications'] = course.applications_count
            course_data['status'] = 'approved'  # Can be extended with moderation
            data.append(course_data)
        
        return Response(data)


class MyJobsView(APIView):
    """Get all jobs owned by the user"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            return Response(
                {"detail": "Доступно только для course_admin и manager."},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        if not user.company:
            return Response([])
        
        if user.role == User.Role.COURSE_ADMIN:
            jobs = JobVacancy.objects.filter(company=user.company, is_active=True)
        else:
            jobs = JobVacancy.objects.filter(company=user.company, is_active=True)
        
        # Add view and application counts
        data = []
        for job in jobs:
            job_data = JobVacancySerializer(job).data
            job_data['views'] = getattr(job, "views", 0)
            job_data['applications'] = getattr(job, "applications_count", 0)
            job_data['status'] = 'approved'
            data.append(job_data)
        
        return Response(data)


class BoostCourseView(APIView):
    """Boost a course (promote to TOP)"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        user = request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            return Response(
                {"detail": "Доступно только для course_admin и manager."},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        course = get_object_or_404(PublicCourse, pk=pk)
        
        # Check ownership
        if user.role == User.Role.COURSE_ADMIN and course.company.owner != user:
            return Response({"detail": "Недостаточно прав."}, status=status.HTTP_403_FORBIDDEN)
        if user.role == User.Role.MANAGER and course.company.owner__company != user.company:
            return Response({"detail": "Недостаточно прав."}, status=status.HTTP_403_FORBIDDEN)
        
        # Check balance
        BOOST_COST = 500
        try:
            company_balance = CompanyBalance.objects.get(company=user.company)
            if company_balance.balance < BOOST_COST:
                return Response(
                    {"detail": f"Недостаточно средств. Требуется {BOOST_COST} eC."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except CompanyBalance.DoesNotExist:
            return Response(
                {"detail": f"Недостаточно средств. Требуется {BOOST_COST} eC."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        # Boost the course
        course.is_promoted = True
        course.promoted_until = timezone.now() + timedelta(days=7)
        course.save()
        
        # Deduct balance
        company_balance.balance -= BOOST_COST
        company_balance.save()
        
        Transaction.objects.create(
            company=user.company,
            amount=-BOOST_COST,
            reason=f"Продвижение курса: {course.title}",
            transaction_type=Transaction.TransactionType.BOOST,
        )
        
        return Response(PublicCourseSerializer(course).data)


class BoostJobView(APIView):
    """Boost a job vacancy (promote to TOP)"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        user = request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            return Response(
                {"detail": "Доступно только для course_admin и manager."},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        job = get_object_or_404(JobVacancy, pk=pk)
        
        # Check ownership
        if user.role == User.Role.COURSE_ADMIN and job.company.owner != user:
            return Response({"detail": "Недостаточно прав."}, status=status.HTTP_403_FORBIDDEN)
        if user.role == User.Role.MANAGER and job.company.owner__company != user.company:
            return Response({"detail": "Недостаточно прав."}, status=status.HTTP_403_FORBIDDEN)
        
        # Check balance
        BOOST_COST = 500
        try:
            company_balance = CompanyBalance.objects.get(company=user.company)
            if company_balance.balance < BOOST_COST:
                return Response(
                    {"detail": f"Недостаточно средств. Требуется {BOOST_COST} eC."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except CompanyBalance.DoesNotExist:
            return Response(
                {"detail": f"Недостаточно средств. Требуется {BOOST_COST} eC."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        # Boost the job
        job.is_promoted = True
        job.promoted_until = timezone.now() + timedelta(days=7)
        job.save()
        
        # Deduct balance
        company_balance.balance -= BOOST_COST
        company_balance.save()
        
        Transaction.objects.create(
            company=user.company,
            amount=-BOOST_COST,
            reason=f"Продвижение вакансии: {job.title}",
            transaction_type=Transaction.TransactionType.BOOST,
        )
        
        return Response(JobVacancySerializer(job).data)


class UrgentCourseView(APIView):
    """Add urgent badge to a course"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        user = request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            return Response(
                {"detail": "Доступно только для course_admin и manager."},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        course = get_object_or_404(PublicCourse, pk=pk)
        
        # Check ownership
        if user.role == User.Role.COURSE_ADMIN and course.company.owner != user:
            return Response({"detail": "Недостаточно прав."}, status=status.HTTP_403_FORBIDDEN)
        if user.role == User.Role.MANAGER and course.company.owner__company != user.company:
            return Response({"detail": "Недостаточно прав."}, status=status.HTTP_403_FORBIDDEN)
        
        # Check balance
        URGENT_COST = 200
        try:
            company_balance = CompanyBalance.objects.get(company=user.company)
            if company_balance.balance < URGENT_COST:
                return Response(
                    {"detail": f"Недостаточно средств. Требуется {URGENT_COST} eC."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except CompanyBalance.DoesNotExist:
            return Response(
                {"detail": f"Недостаточно средств. Требуется {URGENT_COST} eC."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        # Add urgent badge
        course.is_urgent = True
        course.urgent_until = timezone.now() + timedelta(days=3)
        course.save()
        
        # Deduct balance
        company_balance.balance -= URGENT_COST
        company_balance.save()
        
        Transaction.objects.create(
            company=user.company,
            amount=-URGENT_COST,
            reason=f"Срочный бейдж для курса: {course.title}",
            transaction_type=Transaction.TransactionType.URGENT,
        )
        
        return Response(PublicCourseSerializer(course).data)


class UrgentJobView(APIView):
    """Add urgent badge to a job vacancy"""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        user = request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            return Response(
                {"detail": "Доступно только для course_admin и manager."},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        job = get_object_or_404(JobVacancy, pk=pk)
        
        # Check ownership
        if user.role == User.Role.COURSE_ADMIN and job.company.owner != user:
            return Response({"detail": "Недостаточно прав."}, status=status.HTTP_403_FORBIDDEN)
        if user.role == User.Role.MANAGER and job.company.owner__company != user.company:
            return Response({"detail": "Недостаточно прав."}, status=status.HTTP_403_FORBIDDEN)
        
        # Check balance
        URGENT_COST = 200
        try:
            company_balance = CompanyBalance.objects.get(company=user.company)
            if company_balance.balance < URGENT_COST:
                return Response(
                    {"detail": f"Недостаточно средств. Требуется {URGENT_COST} eC."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except CompanyBalance.DoesNotExist:
            return Response(
                {"detail": f"Недостаточно средств. Требуется {URGENT_COST} eC."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        # Add urgent badge
        job.is_urgent = True
        job.urgent_until = timezone.now() + timedelta(days=3)
        job.save()
        
        # Deduct balance
        company_balance.balance -= URGENT_COST
        company_balance.save()
        
        Transaction.objects.create(
            company=user.company,
            amount=-URGENT_COST,
            reason=f"Срочный бейдж для вакансии: {job.title}",
            transaction_type=Transaction.TransactionType.URGENT,
        )
        
        return Response(JobVacancySerializer(job).data)


class PublicCourseViewSet(viewsets.ReadOnlyModelViewSet):
    """Public course listing for marketplace (no auth required)"""
    queryset = PublicCourse.objects.filter(is_active=True).order_by("-is_promoted", "-created_at").select_related("company").prefetch_related("company__landing_pages")
    serializer_class = PublicCourseSerializer
    permission_classes = [permissions.AllowAny]
    lookup_field = "slug"

    def get_serializer_class(self):
        return PublicCourseSerializer

    def get_queryset(self):
        queryset = super().get_queryset().select_related("company").prefetch_related("company__landing_pages")
        
        # Filter by category
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)
        
        # Filter by city
        city = self.request.query_params.get('city')
        if city:
            queryset = queryset.filter(city=city)
        
        # Filter by language in schedule or requirements
        language = self.request.query_params.get('language')
        if language:
            queryset = queryset.filter(
                models.Q(schedule__icontains=language) |
                models.Q(requirements__icontains=language)
            )
        
        # Search by title or description
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(title__icontains=search) |
                models.Q(description__icontains=search)
            )
        
        # Filter by price range
        min_price = self.request.query_params.get('min_price')
        max_price = self.request.query_params.get('max_price')
        if min_price:
            try:
                queryset = queryset.filter(price__gte=float(min_price))
            except (ValueError, TypeError):
                pass
        if max_price:
            try:
                queryset = queryset.filter(price__lte=float(max_price))
            except (ValueError, TypeError):
                pass
        
        return queryset


class PublicJobViewSet(viewsets.ReadOnlyModelViewSet):
    """Public job listing for marketplace (no auth required)"""
    queryset = JobVacancy.objects.filter(is_active=True).order_by("-is_promoted", "-created_at").select_related("company").prefetch_related("company__landing_pages")
    serializer_class = JobVacancySerializer
    permission_classes = [permissions.AllowAny]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            from .serializers import JobVacancyDetailSerializer
            return JobVacancyDetailSerializer
        return JobVacancySerializer

    def get_queryset(self):
        queryset = super().get_queryset().select_related("company").prefetch_related("company__landing_pages")
        
        # Filter by category
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)
        
        # Filter by city
        city = self.request.query_params.get('city')
        if city:
            queryset = queryset.filter(city=city)
        
        # Filter by schedule
        schedule = self.request.query_params.get('schedule')
        if schedule:
            queryset = queryset.filter(schedule__icontains=schedule)
        
        # Search by title or description
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(title__icontains=search) |
                models.Q(description__icontains=search)
            )
        
        # Filter by salary range
        salary_min = self.request.query_params.get('salary_min')
        salary_max = self.request.query_params.get('salary_max')
        if salary_min:
            try:
                queryset = queryset.filter(salary_min__gte=int(salary_min))
            except (ValueError, TypeError):
                pass
        if salary_max:
            try:
                queryset = queryset.filter(salary_max__lte=int(salary_max))
            except (ValueError, TypeError):
                pass
        
        return queryset


# Create your views here.


# ═══════════════════════════════════════════════════════════════════════
#  TELEGRAM BIND CODE GENERATION
# ═══════════════════════════════════════════════════════════════════════

import random


class GenerateTelegramBindCodeView(APIView):
    """
    Generate a one-time code for binding a Telegram account.

    POST /api/bot/generate-bind-code/
    Authenticated user generates a 6-digit code.
    The code is valid for 10 minutes.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user

        # Invalidate any existing pending codes for this user
        TelegramBindCode.objects.filter(
            user=user,
            is_used=False,
            expires_at__gt=timezone.now(),
        ).update(is_used=True)

        # Generate a 6-digit code
        code = f"{random.randint(0, 999999):06d}"
        expires_at = timezone.now() + timedelta(minutes=10)

        bind_code = TelegramBindCode.objects.create(
            user=user,
            code=code,
            expires_at=expires_at,
            is_used=False,
        )

        return Response({
            "code": bind_code.code,
            "expires_at": bind_code.expires_at.isoformat(),
            "message": (
                f"Код действителен 10 минут. "
                f"Используйте в Telegram: /start {user.username} {bind_code.code}"
            ),
        })


class GetTelegramBindCodeView(APIView):
    """
    Get the current active pending bind code for the authenticated user.

    GET /api/bot/bind-code/
    Returns the code if one exists and is still valid.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        pending_code = TelegramBindCode.objects.filter(
            user=user,
            is_used=False,
            expires_at__gt=timezone.now(),
        ).first()

        if not pending_code:
            return Response({
                "code": None,
                "message": "Нет активного кода. Сгенерируйте новый.",
            })

        return Response({
            "code": pending_code.code,
            "expires_at": pending_code.expires_at.isoformat(),
        })


class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.none()
    """
    CRUD для расходов компании.
    """
    permission_classes = [IsCourseAdminOrManager]
    serializer_class = ExpenseSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role == User.Role.COURSE_ADMIN:
            if user.company:
                return Expense.objects.filter(company=user.company)
            return Expense.objects.none()
        if user.role == User.Role.MANAGER:
            if user.company:
                return Expense.objects.filter(company=user.company)
            return Expense.objects.none()
        return Expense.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        company = user.company
        if not company:
            raise PermissionDenied("У вас нет компании для создания расходов.")
        serializer.save(company=company)



class GroupMonthViewSet(viewsets.ModelViewSet):
    queryset = GroupMonth.objects.none()
    """
    CRUD для месяцев обучения группы.
    """
    permission_classes = [IsCourseAdminOrManagerReadOnly]
    serializer_class = GroupMonthSerializer

    def get_queryset(self):
        user = self.request.user
        if user.role in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            if user.company:
                qs = GroupMonth.objects.filter(group__company=user.company)
            else:
                return GroupMonth.objects.none()
        elif user.role == User.Role.TEACHER:
            qs = GroupMonth.objects.filter(group__teacher=user)
        else:
            return GroupMonth.objects.none()
        # Фильтр по группе, если передан параметр ?group=ID
        group_id = self.request.query_params.get("group")
        if group_id and group_id.isdigit():
            qs = qs.filter(group_id=int(group_id))
        # Фильтр по учителю, если передан параметр ?teacher_id=ID
        teacher_id = self.request.query_params.get("teacher_id")
        if teacher_id and teacher_id.isdigit():
            qs = qs.filter(group__teacher_id=int(teacher_id))
        # Фильтр по месяцу, если передан параметр ?month_number=N
        month_number = self.request.query_params.get("month_number")
        if month_number and month_number.isdigit():
            qs = qs.filter(month_number=int(month_number))
        # Фильтр по статусу, если передан параметр ?status=pending|completed
        status = self.request.query_params.get("status")
        if status:
            qs = qs.filter(status=status)
        return qs.select_related("group", "group__course").prefetch_related("group__students")

    def list(self, request, *args, **kwargs):
        # Auto-create все месяцы для группы, если их нет
        group_id = request.query_params.get("group")
        if group_id:
            try:
                group = Group.objects.get(id=group_id)
                if group.total_months and group.total_months > 0:
                    existing = set(GroupMonth.objects.filter(
                        group=group
                    ).values_list("month_number", flat=True))
                    for month_number in range(1, group.total_months + 1):
                        if month_number not in existing:
                            GroupMonth.objects.create(
                                group=group,
                                month_number=month_number,
                                status=GroupMonth.Status.PENDING,
                            )
            except Group.DoesNotExist:
                pass
        return super().list(request, *args, **kwargs)

    def _sync_expense_for_month(self, instance: GroupMonth):
        """
        При завершении месяца с зарплатой — создаёт/обновляет расход (Expense).
        При возврате в ожидание — удаляет расход.
        """
        group = instance.group
        if not group.company:
            return
        desc = f"Зарплата: {group.name} — месяц {instance.month_number}"
        if instance.status == GroupMonth.Status.COMPLETED and instance.teacher_salary:
            Expense.objects.update_or_create(
                company=group.company,
                description=desc,
                defaults={
                    "amount": instance.teacher_salary,
                    "category": "salary",
                    "date": instance.completed_at.date() if instance.completed_at else timezone.localdate(),
                },
            )
        elif instance.status == GroupMonth.Status.PENDING:
            Expense.objects.filter(
                company=group.company,
                description=desc,
            ).delete()

    def perform_update(self, serializer):
        """
        При закрытии месяца (completed_at установлен) —
        авто-создаём следующий месяц, если не превышен лимит total_months.
        """
        instance = serializer.save()

        # Проверяем, был ли месяц только что закрыт
        if instance.status == GroupMonth.Status.COMPLETED and instance.completed_at:
            group = instance.group
            next_month_number = instance.month_number + 1

            # Не создаём, если превышает total_months группы
            if not group.total_months or next_month_number > group.total_months:
                return

            # Создаём следующий месяц, если его ещё нет
            GroupMonth.objects.get_or_create(
                group=group,
                month_number=next_month_number,
                defaults={"status": GroupMonth.Status.PENDING},
            )

            # Автоматическое начисление % учителю при завершении месяца
            self._credit_teacher_percent_on_month_completion(instance)

    def _credit_teacher_percent_on_month_completion(self, instance: GroupMonth):
        """
        При завершении месяца начисляет учителю % от стоимости группы.
        Формула: students_count × course_price × teacher_percent / 100.
        Начисляется один раз за каждый завершённый месяц.
        """
        group = instance.group
        teacher = group.teacher
        teacher_percent = group.teacher_percent or 0
        course = group.course

        if not teacher or not course or not teacher_percent or teacher_percent <= 0:
            return

        student_count = group.students.count() or 0
        if student_count == 0:
            return

        try:
            teacher_user = User.objects.get(id=teacher.id)
            course_price = course.price or 0
            total_amount = course_price * student_count
            percent_amount = int((total_amount * teacher_percent) / 100)

            if percent_amount <= 0:
                return

            # Проверяем, не начислялось ли уже за этот месяц
            # Используем Q-объекты для OR-логики поиска по нескольким условиям
            from django.db.models import Q
            already_credited = UserTransaction.objects.filter(
                user=teacher_user,
                amount__gt=0,
                reason__contains=f"месяц {instance.month_number}",
            ).filter(
                Q(reason__contains=group.name) | Q(reason__contains=str(instance.month_number))
            ).exists()

            if already_credited:
                logger.info(
                    f"Teacher {teacher_user.username} already credited for month {instance.month_number} of group {group.name}"
                )
                return

            teacher_balance, created = UserBalance.objects.get_or_create(user=teacher_user)
            reason = (
                f"Начисление %{teacher_percent}% за месяц {instance.month_number} группы «{group.name}» "
                f"({student_count} студ. × {course_price} eC)"
            )
            teacher_balance.add_coins(percent_amount, reason)
            logger.info(
                f"Teacher {teacher_user.username} credited {percent_amount} eC "
                f"for month {instance.month_number} of group {group.name} "
                f"({teacher_percent}% of {student_count} students × {course_price} eC)"
            )
        except Exception as e:
            logger.warning(f"Failed to credit teacher percent on month completion for group {group.id}: {e}")


class FinanceDashboardView(APIView):
    """
    Дашборд финансовой сводки.
    GET /api/finance/dashboard/
    """
    permission_classes = [IsCourseAdminOrManager]

    def get(self, request):
        from django.db.models import Sum
        
        from calendar import monthrange

        user = request.user
        if user.role == User.Role.COURSE_ADMIN:
            company = user.company
        elif user.role == User.Role.MANAGER:
            company = user.company
        else:
            return Response({'detail': 'Нет доступа'}, status=403)

        if not company:
            return Response({'detail': 'Компания не найдена'}, status=404)

        now = timezone.now().date()
        first_of_month = date(now.year, now.month, 1)
        end_of_month = date(now.year, now.month, monthrange(now.year, now.month)[1])
        start_of_year = date(now.year, 1, 1)

        monthly_income = Payment.objects.filter(
            company=company, status='paid',
            paid_at__gte=first_of_month, paid_at__lte=end_of_month
        ).aggregate(total=Sum('amount'))['total'] or 0

        monthly_expenses = Expense.objects.filter(
            company=company,
            date__gte=first_of_month, date__lte=end_of_month
        ).aggregate(total=Sum('amount'))['total'] or 0

        monthly_salaries = GroupMonth.objects.filter(
            group__company=company, teacher_salary__isnull=False,
            completed_at__gte=first_of_month, completed_at__lte=end_of_month
        ).aggregate(total=Sum('teacher_salary'))['total'] or 0

        monthly_total_expenses = float(monthly_expenses) + float(monthly_salaries)

        yearly_income = Payment.objects.filter(
            company=company, status='paid',
            paid_at__gte=start_of_year, paid_at__lte=end_of_month
        ).aggregate(total=Sum('amount'))['total'] or 0

        yearly_expenses = Expense.objects.filter(
            company=company,
            date__gte=start_of_year, date__lte=end_of_month
        ).aggregate(total=Sum('amount'))['total'] or 0

        yearly_salaries = GroupMonth.objects.filter(
            group__company=company, teacher_salary__isnull=False,
            completed_at__gte=start_of_year, completed_at__lte=end_of_month
        ).aggregate(total=Sum('teacher_salary'))['total'] or 0

        yearly_total_expenses = float(yearly_expenses) + float(yearly_salaries)

        total_debt = Payment.objects.filter(
            company=company, status='debt'
        ).aggregate(total=Sum('amount'))['total'] or 0

        students_count = Student.objects.filter(company=company).count()

        from finance.models import Budget
        active_budgets_count = Budget.objects.filter(company=company, is_active=True).count()

        return Response({
            'monthly': {
                'income': float(monthly_income),
                'expenses': monthly_total_expenses,
                'net': float(monthly_income) - monthly_total_expenses,
            },
            'yearly': {
                'income': float(yearly_income),
                'expenses': yearly_total_expenses,
                'net': float(yearly_income) - yearly_total_expenses,
            },
            'total_debt': float(total_debt),
            'students_count': students_count,
            'active_budgets': active_budgets_count,
        })


class FinanceExportView(APIView):
    """
    Экспорт финансовых данных.
    GET /api/finance/export/<format>/
    Поддерживаемые форматы: json, xlsx, csv
    """
    permission_classes = [IsCourseAdminOrManager]

    def get(self, request, export_format: str):
        from django.db.models import Sum
        from calendar import monthrange
        from django.http import HttpResponse

        user = request.user
        if user.role == User.Role.COURSE_ADMIN:
            company = user.company
        elif user.role == User.Role.MANAGER:
            company = user.company
        else:
            return Response({'detail': 'Нет доступа'}, status=403)

        if not company:
            return Response({'detail': 'Компания не найдена'}, status=404)

        now = timezone.now().date()
        first_of_month = date(now.year, now.month, 1)
        end_of_month = date(now.year, now.month, monthrange(now.year, now.month)[1])

        payments = Payment.objects.filter(
            company=company,
            paid_at__gte=first_of_month, paid_at__lte=end_of_month
        ).select_related('student', 'group')

        expenses = Expense.objects.filter(
            company=company,
            date__gte=first_of_month, date__lte=end_of_month
        )

        if export_format == 'json':
            return Response({
                'payments': list(payments.values('id', 'student__first_name', 'student__last_name', 'amount', 'status', 'paid_at')),
                'expenses': list(expenses.values('id', 'description', 'amount', 'category', 'date')),
            })

        elif export_format == 'xlsx':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()

            # ─── Sheet 1: Payments ───
            ws1 = wb.active
            ws1.title = "Платежи"

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
                top=Side(style="thin", color="E2E8F0"),
                bottom=Side(style="thin", color="E2E8F0"),
            )

            payment_headers = ["ID", "Студент", "Сумма", "Статус", "Дата оплаты"]
            for col_idx, header in enumerate(payment_headers, 1):
                cell = ws1.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_idx, payment in enumerate(payments, 2):
                student_name = f"{payment.student.first_name or ''} {payment.student.last_name or ''}".strip()
                ws1.cell(row=row_idx, column=1, value=payment.id).border = thin_border
                ws1.cell(row=row_idx, column=2, value=student_name or "—").border = thin_border
                ws1.cell(row=row_idx, column=3, value=float(payment.amount)).border = thin_border
                ws1.cell(row=row_idx, column=4, value=payment.get_status_display()).border = thin_border
                ws1.cell(row=row_idx, column=5, value=payment.paid_at.strftime("%d.%m.%Y") if payment.paid_at else "—").border = thin_border

            ws1.column_dimensions["A"].width = 8
            ws1.column_dimensions["B"].width = 35
            ws1.column_dimensions["C"].width = 15
            ws1.column_dimensions["D"].width = 15
            ws1.column_dimensions["E"].width = 15

            # ─── Sheet 2: Expenses ───
            ws2 = wb.create_sheet("Расходы")

            expense_headers = ["ID", "Описание", "Сумма", "Категория", "Дата"]
            for col_idx, header in enumerate(expense_headers, 1):
                cell = ws2.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
                cell.alignment = header_alignment
                cell.border = thin_border

            expense_category_labels = dict(Expense.Category.choices)

            for row_idx, expense in enumerate(expenses, 2):
                ws2.cell(row=row_idx, column=1, value=expense.id).border = thin_border
                ws2.cell(row=row_idx, column=2, value=expense.description).border = thin_border
                ws2.cell(row=row_idx, column=3, value=float(expense.amount)).border = thin_border
                ws2.cell(row=row_idx, column=4, value=expense_category_labels.get(expense.category, expense.category)).border = thin_border
                ws2.cell(row=row_idx, column=5, value=expense.date.strftime("%d.%m.%Y") if expense.date else "—").border = thin_border

            ws2.column_dimensions["A"].width = 8
            ws2.column_dimensions["B"].width = 40
            ws2.column_dimensions["C"].width = 15
            ws2.column_dimensions["D"].width = 25
            ws2.column_dimensions["E"].width = 15

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="finance_{now.strftime("%Y-%m")}.xlsx"'
            wb.save(response)
            return response

        elif export_format == 'csv':
            import csv

            response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
            response["Content-Disposition"] = f'attachment; filename="finance_{now.strftime("%Y-%m")}.csv"'

            writer = csv.writer(response)

            writer.writerow(["=== ПЛАТЕЖИ ==="])
            writer.writerow(["ID", "Студент", "Сумма", "Статус", "Дата оплаты"])
            for payment in payments:
                student_name = f"{payment.student.first_name or ''} {payment.student.last_name or ''}".strip()
                writer.writerow([
                    payment.id,
                    student_name or "—",
                    float(payment.amount),
                    payment.get_status_display(),
                    payment.paid_at.strftime("%d.%m.%Y") if payment.paid_at else "—",
                ])

            writer.writerow([])
            writer.writerow(["=== РАСХОДЫ ==="])
            writer.writerow(["ID", "Описание", "Сумма", "Категория", "Дата"])

            expense_category_labels = dict(Expense.Category.choices)
            for expense in expenses:
                writer.writerow([
                    expense.id,
                    expense.description,
                    float(expense.amount),
                    expense_category_labels.get(expense.category, expense.category),
                    expense.date.strftime("%d.%m.%Y") if expense.date else "—",
                ])

            return response

        else:
            return Response({'detail': f'Формат "{export_format}" не поддерживается. Используйте: json, xlsx, csv'}, status=400)


class UserBalanceMeView(APIView):
    """
    Get current user's eduCoin balance.
    GET /api/user/balance/me/
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        user_balance, created = UserBalance.objects.get_or_create(user=user)
        return Response({'balance': user_balance.balance})


class BroadcastView(APIView):
    """
    Массовая рассылка сообщений через Telegram.
    POST /api/broadcast/send/
    """
    permission_classes = [IsCourseAdminOrManager]

    def post(self, request):
        user = request.user
        company = user.company
        if not company:
            return Response({"detail": "Компания не найдена."}, status=400)

        text = (request.data.get("text") or "").strip()
        target = (request.data.get("target") or "").strip()
        group_id = request.data.get("group_id")

        if not text:
            return Response({"detail": "Текст сообщения обязателен."}, status=400)

        # Санитизация текста через bleach — разрешены только теги Telegram
        text = bleach.clean(
            text,
            tags=['b', 'i', 'u', 's', 'a', 'code', 'pre'],
            attributes={'a': ['href']},
            protocols=['http', 'https', 'mailto'],
            strip=True,
        )

        if len(text) > 4000:
            return Response({"detail": "Текст сообщения не может превышать 4000 символов."}, status=400)

        # Determine recipients
        recipients = []

        if target == "students":
            students = Student.objects.filter(company=company, user__isnull=False)
            for student in students:
                if student.user and student.user.telegram_chat_id:
                    recipients.append({
                        "name": str(student),
                        "chat_id": student.user.telegram_chat_id,
                    })

        elif target == "teachers":
            teachers = User.objects.filter(company=company, role=User.Role.TEACHER)
            for teacher in teachers:
                if teacher.telegram_chat_id:
                    recipients.append({
                        "name": teacher.get_full_name() or teacher.username,
                        "chat_id": teacher.telegram_chat_id,
                    })

        elif target == "managers":
            managers = User.objects.filter(company=company, role=User.Role.MANAGER)
            for mgr in managers:
                if mgr.telegram_chat_id:
                    recipients.append({
                        "name": mgr.get_full_name() or mgr.username,
                        "chat_id": mgr.telegram_chat_id,
                    })

        elif target == "group" and group_id:
            try:
                group = Group.objects.get(id=group_id, company=company)
            except Group.DoesNotExist:
                return Response({"detail": "Группа не найдена."}, status=404)
            for student in group.students.filter(user__isnull=False):
                if student.user and student.user.telegram_chat_id:
                    recipients.append({
                        "name": str(student),
                        "chat_id": student.user.telegram_chat_id,
                    })
        else:
            return Response({"detail": f"Неверный target: {target}"}, status=400)

        if not recipients:
            return Response({"detail": "Нет получателей с привязанным Telegram."}, status=400)

        # Send messages via Telegram
        from telegram_bot.config import _get_application, BOT_TOKEN

        sent_count = 0
        failed_count = 0
        errors = []

        async def send_messages():
            nonlocal sent_count, failed_count
            if not BOT_TOKEN:
                raise Exception("TELEGRAM_BOT_TOKEN не настроен")

            application = _get_application()
            preview_emoji = "📢"
            full_text = f"{preview_emoji} <b>Массовая рассылка</b>\n\n{text}"

            for recipient in recipients:
                try:
                    await application.bot.send_message(
                        chat_id=recipient["chat_id"],
                        text=full_text,
                        parse_mode="HTML",
                    )
                    sent_count += 1
                except Exception as e:
                    failed_count += 1
                    errors.append(f"{recipient['name']}: {str(e)[:100]}")
                    logger.warning(f"Broadcast failed for {recipient['name']}: {e}")

        try:
            from asgiref.sync import async_to_sync
            async_to_sync(send_messages)()
        except Exception as e:
            return Response({"detail": f"Ошибка отправки: {str(e)}"}, status=500)

        return Response({
            "sent": sent_count,
            "failed": failed_count,
            "total": len(recipients),
            "errors": errors[:10],
        })


class ContractViewSet(viewsets.ModelViewSet):
    """
    Договоры / соглашения с PDF-генерацией.
    GET /api/contracts/ — список договоров
    POST /api/contracts/ — создать договор
    GET /api/contracts/{id}/pdf/ — скачать PDF
    """
    queryset = Contract.objects.all().order_by("-created_at")
    serializer_class = ContractSerializer
    permission_classes = [IsCourseAdminOrManager]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.role == User.Role.COURSE_ADMIN:
            return qs.filter(company=user.company)
        if user.role == User.Role.MANAGER:
            if user.company:
                return qs.filter(company=user.company)
            return qs.none()
        return qs.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            raise PermissionDenied("Only course admins and managers can create contracts.")
        student = serializer.validated_data.get("student")
        if student.company != user.company:
            raise PermissionDenied("Student must belong to the same company.")
        group = serializer.validated_data.get("group")
        if group and group.company != user.company:
            raise PermissionDenied("Group must belong to the same company.")
        contract = serializer.save(company=user.company, created_by=user)
        # Auto-generate PDF after creation
        self._generate_contract_pdf(contract)

    def _generate_contract_pdf(self, contract):
        """
        Автоматически генерирует PDF для договора.
        Использует кастомный шаблон компании (ContractTemplate), если он есть,
        иначе — стандартный шаблон core/contract_template.html.
        """
        try:
            from weasyprint import HTML
        except ImportError:
            logger.warning("weasyprint not available, skipping PDF generation")
            return

        if contract.pdf_file:
            return

        company = contract.company
        student = contract.student
        group = contract.group

        admin = User.objects.filter(
            role=User.Role.COURSE_ADMIN, company=company
        ).first()
        admin_name = f"{admin.first_name} {admin.last_name}".strip() or (admin.username if admin else "Администратор")
        course_name = group.course.title if group and group.course else "—"

        context = {
            "contract_number": contract.contract_number,
            "company_name": company.name,
            "company_city": company.city or "Бишкек",
            "company_address": company.district or company.city or "",
            "company_phone": company.phone or "",
            "admin_name": admin_name,
            "student_name": str(student),
            "student_phone": student.phone,
            "student_passport_info": "паспорт: данные паспорта",
            "course_name": course_name,
            "group_name": group.name if group else "—",
            "amount": int(contract.amount),
            "start_date": contract.start_date.strftime("%d.%m.%Y"),
            "end_date": contract.end_date.strftime("%d.%m.%Y") if contract.end_date else "—",
            "payment_terms": "100% предоплата",
            "terms": contract.terms,
        }

        custom_template = ContractTemplate.objects.filter(
            company=company, is_default=True
        ).first()

        if custom_template and custom_template.html_content:
            from django.template import Template, Context
            template = Template(custom_template.html_content)
            html_string = template.render(Context(context))
        else:
            html_string = render_to_string("core/contract_template.html", context)

        try:
            pdf_bytes = HTML(string=html_string).write_pdf()
            contract.pdf_file.save(
                f"contract_{contract.contract_number}.pdf",
                io.BytesIO(pdf_bytes),
                save=True,
            )
            logger.info(f"PDF auto-generated for contract {contract.contract_number}")
        except Exception as e:
            logger.error(f"PDF auto-generation failed for contract {contract.contract_number}: {e}")

    def perform_update(self, serializer):
        user = self.request.user
        instance = self.get_object()
        if instance.company != user.company:
            raise PermissionDenied("Not allowed for this contract.")
        if instance.status != Contract.Status.DRAFT:
            raise PermissionDenied("Only draft contracts can be edited.")
        serializer.save()

    @action(detail=True, methods=["post"])
    def send(self, request, pk=None):
        """Отправить договор студенту (сменить статус на SENT)"""
        contract = self.get_object()
        if contract.company != request.user.company:
            raise PermissionDenied("Not allowed for this contract.")
        if contract.status != Contract.Status.DRAFT:
            return Response({"detail": "Договор уже отправлен."}, status=400)
        contract.status = Contract.Status.SENT
        contract.save(update_fields=["status"])
        return Response(ContractSerializer(contract).data)

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def sign(self, request, pk=None):
        """Подписать договор студентом (сменить статус на SIGNED)"""
        # Используем get_object_or_404 вместо self.get_object(),
        # т.к. get_queryset() фильтрует по company и возвращает qs.none() для студентов
        contract = get_object_or_404(Contract, pk=pk)
        # Only the student who the contract belongs to can sign
        student = getattr(request.user, "student_profile", None)
        if not student or contract.student.id != student.id:
            raise PermissionDenied("Вы не можете подписать этот договор.")
        if contract.status != Contract.Status.SENT:
            return Response({"detail": "Можно подписать только отправленный договор."}, status=400)
        contract.status = Contract.Status.SIGNED
        contract.signed_at = timezone.now()
        contract.save(update_fields=["status", "signed_at"])
        return Response(ContractSerializer(contract).data)
    @action(detail=False, methods=["get", "post"], url_path="auto-fill")
    def auto_fill(self, request):
        """
        Авто-заполнение данных договора на основе студента и группы.
        GET: принимает ?student_id=X&group_id=Y, возвращает предзаполненные данные.
        POST: принимает student_id + group_id, создаёт договор с авто-заполнением.
        """
        from datetime import date

        student_id = request.data.get("student_id") or request.query_params.get("student_id")
        group_id = request.data.get("group_id") or request.query_params.get("group_id")

        if not student_id or not group_id:
            return Response(
                {"detail": "student_id и group_id обязательны."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            student = Student.objects.get(id=student_id, company=request.user.company)
            group = Group.objects.get(id=group_id, company=request.user.company)
        except (Student.DoesNotExist, Group.DoesNotExist):
            raise PermissionDenied("Студент или группа не найдены.")

        course = group.course
        course_name = course.title if course else "—"
        amount = course.price if course else 0

        data = {
            "student": student.id,
            "student_name": str(student),
            "student_phone": student.phone,
            "group": group.id,
            "group_name": group.name,
            "course_name": course_name,
            "amount": float(amount),
            "start_date": group.start_date.isoformat() if group.start_date else date.today().isoformat(),
            "end_date": group.end_date.isoformat() if group.end_date else "",
        }

        if request.method == "POST":
            # Create contract with auto-filled data
            serializer = self.get_serializer(data={
                "student": student.id,
                "group": group.id,
                "amount": amount,
                "start_date": group.start_date or date.today(),
                "end_date": group.end_date,
            })
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

        return Response(data)



    @action(detail=True, methods=["get"], url_path="pdf")
    def download_pdf(self, request, pk=None):
        """Сгенерировать и скачать PDF договора"""
        contract = self.get_object()
        if contract.company != request.user.company:
            raise PermissionDenied("Not allowed for this contract.")

        try:
            from weasyprint import HTML
        except ImportError:
            return Response(
                {"detail": "PDF generation is not available."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # Prepare template context
        company = contract.company
        admin = User.objects.filter(
            role=User.Role.COURSE_ADMIN, company=company
        ).first()
        admin_name = f"{admin.first_name} {admin.last_name}".strip() or (admin.username if admin else "Администратор")
        student = contract.student
        group = contract.group
        course_name = group.course.title if group and group.course else "—"

        context = {
            "contract_number": contract.contract_number,
            "company_name": company.name,
            "company_city": company.city or "Бишкек",
            "company_address": company.district or company.city or "",
            "company_phone": company.phone or "",
            "admin_name": admin_name,
            "student_name": str(student),
            "student_phone": student.phone,
            "student_passport_info": "паспорт: данные паспорта",
            "course_name": course_name,
            "group_name": group.name if group else "—",
            "amount": int(contract.amount),
            "start_date": contract.start_date.strftime("%d.%m.%Y"),
            "end_date": contract.end_date.strftime("%d.%m.%Y") if contract.end_date else "—",
            "payment_terms": "100% предоплата",
            "terms": contract.terms,
        }

        html_string = render_to_string("core/contract_template.html", context)

        try:
            pdf_file = HTML(string=html_string).write_pdf()
            response = HttpResponse(pdf_file, content_type="application/pdf")
            filename = f"contract_{contract.contract_number}.pdf"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            return Response(
                {"detail": f"Ошибка генерации PDF: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class ContractTemplateViewSet(viewsets.ModelViewSet):
    """
    Шаблоны договоров для компании.
    GET /api/contract-templates/ — список шаблонов
    POST /api/contract-templates/ — создать шаблон
    GET /api/contract-templates/{id}/ — детали шаблона
    PUT /api/contract-templates/{id}/ — обновить шаблон
    DELETE /api/contract-templates/{id}/ — удалить шаблон
    """
    queryset = ContractTemplate.objects.all().order_by("-is_default", "-created_at")
    serializer_class = ContractTemplateSerializer
    permission_classes = [IsCourseAdminOrManager]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.role == User.Role.COURSE_ADMIN:
            return qs.filter(company=user.company)
        if user.role == User.Role.MANAGER and user.company:
            return qs.filter(company=user.company)
        return qs.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role not in (User.Role.COURSE_ADMIN, User.Role.MANAGER):
            raise PermissionDenied("Only course admins and managers can create contract templates.")
        # If this template is set as default, unset others
        if serializer.validated_data.get("is_default"):
            ContractTemplate.objects.filter(company=user.company).update(is_default=False)
        serializer.save(company=user.company)

    def perform_update(self, serializer):
        user = self.request.user
        instance = self.get_object()
        if instance.company != user.company:
            raise PermissionDenied("Not allowed for this template.")
        if serializer.validated_data.get("is_default"):
            ContractTemplate.objects.filter(company=user.company).exclude(id=instance.id).update(is_default=False)
        serializer.save()


class StudentContractsView(APIView):
    """
    Список договоров для студента (личный кабинет).
    GET /api/auth/student/contracts/
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.role != User.Role.STUDENT:
            raise PermissionDenied("Only students can access their contracts.")
        student = getattr(request.user, "student_profile", None)
        if not student:
            return Response({"detail": "Student profile not found."}, status=status.HTTP_404_NOT_FOUND)

        contracts = Contract.objects.filter(student=student).order_by("-created_at")
        serializer = ContractSerializer(contracts, many=True, context={"request": request})
        return Response(serializer.data)



class CspReportView(APIView):
    """
    Public endpoint для сбора CSP violation report-ов.
    Браузеры отправляют POST с Content-Type application/csp-report (не application/json!), 
    поэтому читаем тело вручную через json.loads(request.body).
    Все нарушения логируются для мониторинга.
    """
    permission_classes = [permissions.AllowAny]
    authentication_classes = []  # No auth needed for CSP reports

    def post(self, request):
        import json
        try:
            report = json.loads(request.body)
        except (ValueError, AttributeError, TypeError):
            # Невалидный JSON или пустое тело — игнорируем
            return Response(status=204)
        
        # CSP report может быть в формате {"csp-report": {...}} или плоским
        csp_report = report.get("csp-report", report)
        
        if not isinstance(csp_report, dict):
            return Response(status=204)
        
        blocked_uri = csp_report.get("blocked-uri", "unknown")
        violated_directive = csp_report.get("violated-directive", "unknown")
        document_uri = csp_report.get("document-uri", "unknown")
        original_policy = csp_report.get("original-policy", "")
        disposition = csp_report.get("disposition", "unknown")
        source_file = csp_report.get("source-file", "")
        line_number = csp_report.get("line-number", "")
        
        logger.warning(
            "CSP Violation | directive=%s | blocked=%s | document=%s | disposition=%s | source=%s:%s | policy=%s",
            violated_directive,
            blocked_uri,
            document_uri,
            disposition,
            source_file,
            line_number,
            original_policy[:500],
        )
        
        # В production можно также сохранять в БД, отправлять в Sentry и т.д.
        if settings.DEBUG:
            logger.debug("Full CSP report: %s", csp_report)
        
        return Response(status=204)  # No content — браузеру не нужен ответ

